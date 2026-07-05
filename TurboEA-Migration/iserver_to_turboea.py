#!/usr/bin/env python3
"""
iServer -> Turbo-EA workspace-transfer bundle builder
=====================================================

Reads iServer table exports (JSON) and produces a Turbo-EA "Workspace Transfer"
bundle:  <out>.zip  =  manifest.json + workspace.xlsx

The bundle is uploaded at:  Turbo-EA  ->  /admin/settings?tab=migration
(Workspace Transfer).  For a clean migration, import into a *fresh* workspace so
there are zero conflicts.

WHY A TEMPLATE IS NEEDED
------------------------
Turbo-EA's workspace.xlsx has a fixed set of sheets. The metamodel sheets
(CardTypes, RelationTypes, Roles, StakeholderRoles, ResourceTypes,
ComplianceRegs, Settings) are the built-in Turbo-EA model and are identical
across instances. We reuse them from ANY Turbo-EA export used as a template.
This script only *fills* the Cards + Relations sheets (and, optionally, Users)
and clears every other content sheet.

USAGE
-----
    python iserver_to_turboea.py \
        --exports  "F:\\Iserver\\Exports" \
        --template "F:\\Iserver\\TurboEA-Migration\\template-workspace.xlsx" \
        --out      "F:\\Iserver\\Exports\\iserver-real-workspace.zip" \
        [--attributes] [--users]

Run with no args to use the defaults in CONFIG below.

ON REAL DATA - READ THIS
------------------------
1. Extend TYPE_MAP.  Your real repository may use iServer object types not in
   the map below. After a run, open  unmapped_types.txt  (written next to the
   bundle): it lists every skipped iServer type with a count. Add the ones you
   want to TYPE_MAP -> re-run.
2. Encoding.  If names come out as "??????" (mojibake), the JSON export step
   lost UTF-8. Re-export with UTF-8 (see export_iserver_json.py / README).
3. --attributes carries iServer attribute values into each card's `attributes`
   JSON. Turbo-EA only *renders* keys defined in that card type's field schema;
   unknown keys are carried but may not display. Off by default.
4. --users maps iServer users (those with an email) into the Users sheet.
   Roles default to 'member' - promote admins in Turbo-EA after import.
"""

import argparse
import datetime
import json
import os
import sys
import zipfile
from collections import Counter, defaultdict

import openpyxl

# --------------------------------------------------------------------------- #
# CONFIG - defaults; override with CLI flags
# --------------------------------------------------------------------------- #
HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG = {
    "exports":  os.path.join(os.path.dirname(HERE), "Exports"),
    "template": os.path.join(HERE, "template-workspace.xlsx"),
    "out":      os.path.join(os.path.dirname(HERE), "Exports", "iserver-real-workspace.zip"),
}

# iServer ExactObjectType name  ->  Turbo-EA CardType key.
# Turbo-EA's 13 built-in card types:
#   Objective, Platform, Initiative, Organization, Interface, BusinessCapability,
#   BusinessContext, BusinessProcess, Application, DataObject, ITComponent,
#   TechCategory, Provider
# Extend this table for your real repository (see unmapped_types.txt after a run).
TYPE_MAP = {
    "Process":                         "BusinessProcess",
    "Task":                            "BusinessProcess",
    "Organization Unit":               "Organization",
    "Actor":                           "Organization",
    "Role":                            "Organization",
    "Capability":                      "BusinessCapability",
    "Function":                        "BusinessCapability",
    "Business Service":                "BusinessCapability",
    "Physical Application Component":  "Application",
    "Logical Application Component":   "Application",
    "Physical Data Component":         "DataObject",
    "Logical Data Component":          "DataObject",
    "Data Entity":                     "DataObject",
    "Physical Technology Component":   "ITComponent",
    "Logical Technology Component":    "ITComponent",
    "Technology Service":              "ITComponent",
    "Interface":                       "Interface",
}

# Sheets carried over from the template unchanged (built-in Turbo-EA metamodel).
# Everything NOT listed here and NOT filled by this script is cleared to header-only.
STRUCTURE_SHEETS = {
    "CardTypes", "RelationTypes", "Roles", "StakeholderRoles",
    "ResourceTypes", "ComplianceRegs", "Settings",
}


# --------------------------------------------------------------------------- #
def load_json(path):
    with open(path, encoding="utf-8", errors="replace") as f:
        return json.load(f)


def truthy(v):
    return str(v).strip().lower() == "true"


def build_cards(objs, id2typename, obj_attrs):
    """Return (card_rows, objid2card, unmapped_counter)."""
    objid2card = {}
    used = defaultdict(set)
    rows = []
    unmapped = Counter()
    for o in objs:
        if truthy(o.get("DeleteFlag")):
            continue
        exact = id2typename.get(o.get("ExactObjectTypeID"))
        tea = TYPE_MAP.get(exact)
        if not tea:
            unmapped[exact] += 1
            continue
        name = (o.get("ObjectName") or "").strip()
        if not name:
            continue
        final = name
        if final in used[tea]:
            i = 2
            while f"{name} ({i})" in used[tea]:
                i += 1
            final = f"{name} ({i})"
        used[tea].add(final)
        oid = o.get("ObjectID")
        objid2card[oid] = (tea, final)
        attrs = obj_attrs.get(oid)
        rows.append({
            "type": tea,
            "name": final,
            "parent_path": None,
            "subtype": None,
            "description": (o.get("ObjectDescription") or "").strip() or None,
            "external_id": oid,
            "alias": None,
            "approval_status": "APPROVED",
            "status": "ACTIVE",
            "source_system": "iServer",
            "confidence": None,
            "lifecycle": None,
            "attributes": json.dumps(attrs, ensure_ascii=False) if attrs else None,
        })
    return rows, objid2card, unmapped


def build_pair_map(ws_reltypes):
    """(source_type_key, target_type_key) -> relation key. Prefer non-successor."""
    hdr = [c.value for c in ws_reltypes[1]]
    pair2key, deferred = {}, []
    for row in ws_reltypes.iter_rows(min_row=2, values_only=True):
        d = dict(zip(hdr, row))
        pair = (d["source_type_key"], d["target_type_key"])
        if "Successor" in (d["key"] or ""):
            deferred.append((pair, d["key"]))
        else:
            pair2key.setdefault(pair, d["key"])
    for pair, key in deferred:
        pair2key.setdefault(pair, key)
    return pair2key


def build_relations(rels, objid2card, pair2key):
    """Return (relation_rows, skip_counter). Only valid, de-duplicated pairs."""
    rows, seen, skips = [], set(), Counter()
    for r in rels:
        f = objid2card.get(r.get("FromObjectId"))
        t = objid2card.get(r.get("ToObjectId"))
        if not f or not t:
            skips["endpoint_not_mapped"] += 1
            continue
        if (f[0], t[0]) in pair2key:
            src, tgt, key = f, t, pair2key[(f[0], t[0])]
        elif (t[0], f[0]) in pair2key:
            src, tgt, key = t, f, pair2key[(t[0], f[0])]
        else:
            skips[f"no_reltype:{f[0]}->{t[0]}"] += 1
            continue
        dedup = (key, src[0], src[1], tgt[0], tgt[1])
        if dedup in seen:
            skips["duplicate"] += 1
            continue
        seen.add(dedup)
        rows.append({
            "type": key,
            "source_type": src[0], "source_ref": src[1],
            "target_type": tgt[0], "target_ref": tgt[1],
            "description": None, "attributes": None,
        })
    return rows, skips


def build_obj_attrs(exports):
    """ObjectId -> {attr_name: value}. Empty dict if attribute files absent."""
    apath = os.path.join(exports, "Attribute.json")
    vpath = os.path.join(exports, "AttributeValue.json")
    if not (os.path.exists(apath) and os.path.exists(vpath)):
        print("  ! attribute files not found - skipping attribute enrichment")
        return {}
    id2name = {a["AttributeId"]: a["AttributeName"] for a in load_json(apath)}
    out = defaultdict(dict)
    for v in load_json(vpath):
        name = id2name.get(v.get("AttributeId"))
        if not name:
            continue
        val = next((v[k] for k in
                    ("ValueText", "ValueFloat", "ValueBigInt", "ValueDate")
                    if v.get(k) not in (None, "")), None)
        if val is None:
            continue
        out[v["ObjectId"]][name] = val
    return out


def build_users(exports):
    """iServer users with an email -> Users sheet rows."""
    upath = os.path.join(exports, "User.json")
    if not os.path.exists(upath):
        return []
    rows, seen = [], set()
    for u in load_json(upath):
        if truthy(u.get("DeleteFlag")):
            continue
        email = (u.get("UserEmail") or u.get("BusinessEmail") or "").strip()
        if not email or email.lower() in seen:
            continue
        seen.add(email.lower())
        sso = truthy(u.get("IsActiveDirectoryUser")) or truthy(u.get("IsSsoAuthenticated"))
        rows.append({
            "email": email,
            "display_name": (u.get("Username") or email).strip(),
            "role": "member",
            "is_active": True,
            "auth_provider": "sso" if sso else "local",
            "locale": "en",
        })
    return rows


def fill_sheet(ws, rows):
    hdr = [c.value for c in ws[1]]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for rd in rows:
        ws.append([rd.get(h) for h in hdr])


def main():
    ap = argparse.ArgumentParser(description="Build a Turbo-EA workspace bundle from iServer JSON exports.")
    ap.add_argument("--exports",  default=CONFIG["exports"],  help="Folder with iServer *.json exports")
    ap.add_argument("--template", default=CONFIG["template"], help="A Turbo-EA workspace.xlsx used as metamodel template")
    ap.add_argument("--out",      default=CONFIG["out"],      help="Output .zip bundle path")
    ap.add_argument("--attributes", action="store_true", help="Carry iServer attribute values into cards")
    ap.add_argument("--users",      action="store_true", help="Map iServer users into the Users sheet")
    args = ap.parse_args()

    for p in (args.exports, args.template):
        if not os.path.exists(p):
            sys.exit(f"ERROR: not found: {p}")

    print("Loading iServer exports ...")
    objs  = load_json(os.path.join(args.exports, "Object.json"))
    types = load_json(os.path.join(args.exports, "ObjectType.json"))
    rels  = load_json(os.path.join(args.exports, "Relation.json"))
    id2typename = {t["ObjectTypeID"]: t["ObjectTypeName"] for t in types}

    obj_attrs = build_obj_attrs(args.exports) if args.attributes else {}

    print("Mapping objects -> cards ...")
    card_rows, objid2card, unmapped = build_cards(objs, id2typename, obj_attrs)

    print("Mapping relations ...")
    wb = openpyxl.load_workbook(args.template)
    pair2key = build_pair_map(wb["RelationTypes"])
    rel_rows, skips = build_relations(rels, objid2card, pair2key)

    user_rows = build_users(args.exports) if args.users else []

    # ---- write workbook: fill Cards/Relations(/Users), keep structure, clear rest
    filled = {"Cards": card_rows, "Relations": rel_rows}
    if args.users:
        filled["Users"] = user_rows
    for ws in wb.worksheets:
        if ws.title in filled:
            fill_sheet(ws, filled[ws.title])
        elif ws.title not in STRUCTURE_SHEETS:
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    tmp_xlsx = os.path.join(os.path.dirname(os.path.abspath(args.out)), "_workspace.xlsx")
    wb.save(tmp_xlsx)

    # ---- manifest (counts must match the sheets)
    chk = openpyxl.load_workbook(tmp_xlsx, read_only=True)
    sections = {ws.title: max((ws.max_row or 1) - 1, 0) for ws in chk.worksheets}
    chk.close()
    manifest = {
        "format_version": "1",
        "app_version": "1.62.4",
        "exported_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "source_url": "",
        "include_archived": True,
        "sections": sections,
        "assets": [],
    }

    if os.path.exists(args.out):
        os.remove(args.out)
    with zipfile.ZipFile(args.out, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("manifest.json", json.dumps(manifest, indent=2))
        z.write(tmp_xlsx, "workspace.xlsx")
    os.remove(tmp_xlsx)

    # ---- reports
    rep = os.path.join(os.path.dirname(os.path.abspath(args.out)), "unmapped_types.txt")
    with open(rep, "w", encoding="utf-8") as f:
        f.write("iServer object types NOT mapped (add wanted ones to TYPE_MAP):\n\n")
        for name, c in unmapped.most_common():
            f.write(f"{c:8}  {name}\n")

    print("\n" + "=" * 60)
    print(f"Cards:     {len(card_rows)}")
    for k, v in Counter(c['type'] for c in card_rows).most_common():
        print(f"             {v:6}  {k}")
    print(f"Relations: {len(rel_rows)}")
    if args.users:
        print(f"Users:     {len(user_rows)}")
    print(f"Bundle:    {args.out}  ({os.path.getsize(args.out)/1024:.1f} KB)")
    print(f"Unmapped types report: {rep}")
    print("=" * 60)
    print("Import at Turbo-EA -> /admin/settings?tab=migration (Workspace Transfer)")
    print("Use a FRESH workspace for a zero-conflict import.")


if __name__ == "__main__":
    main()
