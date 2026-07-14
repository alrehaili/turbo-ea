"""End-to-end stage → apply test for the NORA template adapter ([FORK] WP6.6).

Builds a synthetic بنية الأعمال workbook, runs it through the generic
staging + apply pipeline, and asserts the landed cards, hierarchy, subtype
routing, name-stub behaviour and relations. Requires the DB conftest.
"""

from __future__ import annotations

import pytest
from openpyxl import Workbook  # type: ignore[import-untyped]
from sqlalchemy import select

from app.models.card import Card
from app.models.migration import Migration
from app.models.relation import Relation
from app.services.migration.apply import apply_migration
from app.services.migration.registry import get_source
from app.services.migration.staging import stage_cards, stage_relations
from tests.conftest import create_card, create_card_type, create_relation_type, create_user

pytestmark = pytest.mark.asyncio


def _write_sheet(wb: Workbook, name: str, rows: list[list]) -> None:
    if wb.sheetnames == ["Sheet"]:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)
    for row in rows:
        ws.append(row)


def _business_workbook_path(tmp_path) -> str:
    wb = Workbook()
    _write_sheet(
        wb,
        "دليل الخدمات",
        [
            ["دليل الخدمات"],
            [
                "رمز الخدمة",
                "اسم الخدمة",
                "وصف الخدمة",
                "تصنيف الخدمة",
                "الخدمة الأساسية",
                "مستوى أتمتة الخدمة",
                "التطبيقات المستخدمة لتنفيذ الخدمة",
            ],
            ["شرح"] * 7,
            [
                "S02",
                "خدمة تقديم شكوى",
                "استلام شكاوى المستفيدين",
                "فرعي",
                "خدمة العناية بالمستفيدين",
                "مؤتمت جزئيا",
                "نظام إدارة الشكاوى",
            ],
            [
                "S01",
                "خدمة العناية بالمستفيدين",
                "الخدمة الرئيسية للعناية بالمستفيدين",
                "رئيسي",
                None,
                "مؤتمت كليا",
                None,
            ],
        ],
    )
    p = tmp_path / "business.xlsx"
    wb.save(p)
    return str(p)


async def _setup_metamodel(db) -> None:
    await create_card_type(db, key="GovService", label="Government Service", has_hierarchy=True)
    await create_card_type(db, key="Application", label="Application")
    await create_relation_type(
        db,
        key="relGovServiceToApp",
        label="is delivered by",
        source_type_key="GovService",
        target_type_key="Application",
    )


async def test_business_workbook_stages_and_applies(db, tmp_path):
    await _setup_metamodel(db)
    # A pre-existing card with the referenced application's name — the stub
    # must bind to it (name+type fallback) instead of duplicating, and must
    # not blank its description.
    user = await create_user(db, email="importer@test.com", role="admin")
    existing_app = await create_card(
        db,
        card_type="Application",
        name="نظام إدارة الشكاوى",
        user_id=user.id,
        description="نظام قائم",
    )

    source = get_source("nora_xlsx")
    snapshot = source.parse(_business_workbook_path(tmp_path))
    assert not [e for e in snapshot.parse_errors if "no name column" in e]

    migration = Migration(
        name="business.xlsx",
        source_type="nora_xlsx",
        file_hash="b" * 64,
        status="parsed",
    )
    db.add(migration)
    await db.flush()

    card_stats = await stage_cards(db, migration, source, snapshot)
    rel_stats = await stage_relations(db, migration, source, snapshot)
    # Two service rows + one referenced-app stub; the stub resolves to the
    # existing card so it stages as update/skip, not create.
    assert card_stats["create"] == 2
    assert card_stats["update"] + card_stats["skip"] == 1
    assert rel_stats["create"] == 1
    assert rel_stats["conflict"] == 0

    counts = await apply_migration(db, migration, user)
    assert counts["errors"] == 0

    sub = (
        await db.execute(
            select(Card).where(Card.type == "GovService", Card.name == "خدمة تقديم شكوى")
        )
    ).scalar_one()
    main = (
        await db.execute(
            select(Card).where(Card.type == "GovService", Card.name == "خدمة العناية بالمستفيدين")
        )
    ).scalar_one()
    assert sub.parent_id == main.id
    assert sub.attributes["serviceClassification"] == "sub"
    assert sub.attributes["automationLevel"] == "partiallyAutomated"
    assert main.attributes["serviceClassification"] == "main"

    # The stub bound to the pre-existing application — no duplicate, and the
    # description survived.
    apps = (await db.execute(select(Card).where(Card.type == "Application"))).scalars().all()
    assert [a.id for a in apps] == [existing_app.id]
    assert apps[0].description == "نظام قائم"

    rel = (
        await db.execute(select(Relation).where(Relation.type == "relGovServiceToApp"))
    ).scalar_one()
    assert rel.source_id == sub.id
    assert rel.target_id == existing_app.id

    # Re-running the same workbook is a no-op (idempotent by identity map).
    migration2 = Migration(
        name="business2.xlsx", source_type="nora_xlsx", file_hash="c" * 64, status="parsed"
    )
    db.add(migration2)
    await db.flush()
    stats2 = await stage_cards(db, migration2, source, snapshot)
    assert stats2["create"] == 0
    rel_stats2 = await stage_relations(db, migration2, source, snapshot)
    assert rel_stats2["create"] == 0
    assert rel_stats2["skip"] == 1


async def test_technology_sheet_routes_subtypes(db, tmp_path):
    await create_card_type(db, key="ITComponent", label="IT Component")
    user = await create_user(db, email="importer2@test.com", role="admin")

    wb = Workbook()
    _write_sheet(
        wb,
        "أجهزة المضيف المادي",
        [
            ["دليل أجهزة المضيف المادي"],
            ["اسم المضيف", "الشركة المصنعة", "حالة عقد الدعم مع المزود"],
            ["شرح"] * 3,
            ["RYD-Host01", "DELL", "Active"],
        ],
    )
    path = tmp_path / "tech.xlsx"
    wb.save(path)

    source = get_source("nora_xlsx")
    snapshot = source.parse(str(path))
    migration = Migration(
        name="tech.xlsx", source_type="nora_xlsx", file_hash="d" * 64, status="parsed"
    )
    db.add(migration)
    await db.flush()

    await stage_cards(db, migration, source, snapshot)
    counts = await apply_migration(db, migration, user)
    assert counts["errors"] == 0

    host = (
        await db.execute(select(Card).where(Card.type == "ITComponent", Card.name == "RYD-Host01"))
    ).scalar_one()
    assert host.subtype == "physicalHost"
    assert host.attributes["manufacturer"] == "DELL"
    assert host.attributes["supportContractStatus"] == "active"
