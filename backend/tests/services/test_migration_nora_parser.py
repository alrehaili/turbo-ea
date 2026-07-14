"""Unit tests for the NORA/DGA data-collection template parser ([FORK] WP6.6).

Synthetic workbooks built in-memory mirror the official حصر البيانات layout
(row 1 title, row 2 Arabic headers, row 3 explanations, row 4+ data) —
never real agency data.
"""

from __future__ import annotations

from openpyxl import Workbook  # type: ignore[import-untyped]

from app.services.migration.registry import SOURCES
from app.services.migration.sources.nora_xlsx.xlsx_parser import (
    is_xlsx_payload,
    parse_xlsx_path,
)


def _write_sheet(wb: Workbook, name: str, rows: list[list]) -> None:
    if wb.sheetnames == ["Sheet"]:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)
    for row in rows:
        ws.append(row)


def _save(wb: Workbook, tmp_path) -> str:
    p = tmp_path / "template.xlsx"
    wb.save(p)
    return str(p)


def _by_id(snapshot):
    return {e.source_id: e for e in snapshot.entities}


# ---------------------------------------------------------------------------
# Business architecture workbook
# ---------------------------------------------------------------------------


def _business_workbook() -> Workbook:
    wb = Workbook()
    _write_sheet(wb, "المقدمة", [["مقدمة"]])
    _write_sheet(
        wb,
        "دليل الخدمات",
        [
            ["دليل الخدمات"],
            [
                "رمز الخدمة",
                "اسم الخدمة",
                "وصف الخدمة",
                "تصنيف الخدمة",
                "الخدمة الأساسية",
                "نوع الخدمة",
                "مستوى أتمتة الخدمة",
                "التطبيقات المستخدمة لتنفيذ الخدمة",
                "التغطية الجغرافية للخدمة",
                "قنوات تقديم الخدمة",
                "نوع المستفيد من الخدمة",
                "",
                "",
                "مستوى نضج الخدمة",
                "وجود رسوم الخدمة",
                "معدل طلبات الخدمة شهريا",
                "اتفاقية مستوى الخدمة",
            ],
            [
                "شرح",
                "شرح",
                "شرح",
                "شرح",
                "شرح",
                "شرح",
                "شرح",
                "شرح",
                "شرح",
                "شرح",
                "أفراد",
                "قطاع أعمال",
                "جهات حكومية",
                "شرح",
                "شرح",
                "شرح",
                "شرح",
            ],
            [
                "S02",
                "خدمة تقديم شكوى",
                "استلام شكاوى المستفيدين",
                "فرعي",
                "خدمة العناية بالمستفيدين",
                "داعمة",
                "مؤتمت جزئيا",
                "نظام إدارة الشكاوى\nبوابة الجهة",
                "كافة مناطق المملكة",
                "•مركز الاتصال بالجهة\n•الموقع الإلكتروني للجهة\n•تطبيق الأجهزة الذكية",
                "نعم",
                "نعم",
                "لا",
                "تفاعلية",
                "لا",
                "7500 مرة",
                "10 أيام عمل",
            ],
        ],
    )
    _write_sheet(
        wb,
        "دليل الإجراءات",
        [
            ["دليل الإجراءات"],
            [
                "رمز الإجراء",
                "اسم الإجراء",
                "وصف الإجراء",
                "تصنيف الإجراء",
                "مستوى أتمتة الإجراء",
                "الأنظمة المستخدمة لتنفيذ الإجراء",
                "خدمات ذات علاقة",
                "مدة تنفيذ إجراء العمل",
                "حدث الإطلاق",
                "قواعد العمل",
            ],
            ["شرح"] * 10,
            [
                "P01",
                "معالجة شكاوى المستفيدين",
                "التعامل مع الشكاوى",
                "رئيسي",
                "مؤتمت جزئيا",
                "نظام إدارة الشكاوى",
                "خدمة تقديم شكوى",
                "14 يوم عمل",
                "استقبال شكوى",
                "تصنيف الشكاوى حسب طبيعتها",
            ],
        ],
    )
    _write_sheet(wb, "Lookups", [["رئيسي"]])
    return wb


def test_business_workbook_services_and_processes(tmp_path):
    snapshot = parse_xlsx_path(_save(_business_workbook(), tmp_path))
    ents = _by_id(snapshot)

    svc = ents["nora:GovService:خدمه تقديم شكوي"]
    assert svc.type == "GovService"
    assert svc.name == "خدمة تقديم شكوى"
    assert svc.description == "استلام شكاوى المستفيدين"
    attrs = svc.custom_fields
    assert attrs["serviceCode"] == "S02"
    assert attrs["serviceClassification"] == "sub"
    assert attrs["serviceType"] == "supporting"
    assert attrs["automationLevel"] == "partiallyAutomated"
    assert attrs["geoCoverage"] == "كافة مناطق المملكة"
    assert set(attrs["deliveryChannel"]) == {"callCenter", "portal", "mobileApp"}
    assert attrs["beneficiaryType"] == ["citizen", "business"]
    assert attrs["serviceMaturity"] == "interactive"
    assert attrs["feeModel"] == "free"
    assert attrs["monthlyTransactions"] == 7500
    assert attrs["slaDays"] == 10

    # Parent (main service) becomes a stub + hierarchy edge.
    assert svc.parent_id == "nora:GovService:خدمه العنايه بالمستفيدين"
    parent = ents[svc.parent_id]
    assert parent.raw.get("nora_stub") is True

    proc = ents["nora:BusinessProcess:معالجه شكاوي المستفيدين"]
    assert proc.custom_fields["processClassification"] == "main"
    assert proc.custom_fields["automationLevel"] == "partiallyAutomated"
    assert proc.custom_fields["durationDays"] == 14
    assert proc.custom_fields["triggerEvent"] == "استقبال شكوى"

    # Referenced applications are stubbed once, even across sheets.
    app_stub = ents["nora:Application:نظام اداره الشكاوي"]
    assert app_stub.type == "Application"

    rel_types = {(r.type, r.from_entity_id, r.to_entity_id) for r in snapshot.relations}
    assert (
        "serviceToApplication",
        svc.source_id,
        "nora:Application:نظام اداره الشكاوي",
    ) in rel_types
    assert (
        "serviceToApplication",
        svc.source_id,
        "nora:Application:بوابه الجهه",
    ) in rel_types
    # Procedure's related-service column links service → process.
    assert ("serviceToProcess", svc.source_id, proc.source_id) in rel_types
    assert (
        "processToApplication",
        proc.source_id,
        "nora:Application:نظام اداره الشكاوي",
    ) in rel_types

    # Intro + lookup sheets are noted, not imported.
    assert any("المقدمة" in e for e in snapshot.parse_errors)
    assert any("Lookups" in e for e in snapshot.parse_errors)


# ---------------------------------------------------------------------------
# Applications architecture workbook
# ---------------------------------------------------------------------------


def _applications_workbook() -> Workbook:
    wb = Workbook()
    _write_sheet(
        wb,
        "سجل التطبيقات",
        [
            ["سجل التطبيقات"],
            [
                "الاسم",
                "الوصف",
                "الطبقة",
                "عدد المستخدمين",
                "نوع التطوير",
                "نوع المصدر",
                "المقاول",
                "رابط التطبيق",
                "نوع المصادقة",
                "تاريخ الاطلاق",
                "الأهمية",
                "حالة التطبيق",
                "نمط الهيكلية المعيارية",
                "التكلفة الرأسمالية",
                "التكلفة التشغيلية",
            ],
            ["شرح"] * 15,
            [
                "بوابة الجهة الخارجية",
                "تطبيق صفحات ويب",
                "Access Application Layer – طبقة الوصول",
                12000,
                "COTS – كود مصدري جاهز",
                "In-House Development – تم التطوير داخليا",
                "شركة التقنية",
                "http://www.example.gov.sa",
                "الدخول الموحد",
                "2021-12-12",
                "عالية",
                "Active - نشط",
                "N-Tier Architecture",
                "1,000,000 ريال",
                "500,000 ريال",
            ],
        ],
    )
    _write_sheet(
        wb,
        "نقاط الربط التقني",
        [
            ["سجل نقاط الربط التقني"],
            [
                "الرقم",
                "الاسم",
                "الوصف",
                "المستهلك",
                "المنتج",
                "نطاق الربط",
                "المدخلات",
                "المخرجات",
                "منصة التكامل",
                "نوع الربط",
                "طريقة الربط",
                "صيغة البيانات",
            ],
            ["شرح"] * 12,
            [
                "Ex01",
                "استعلام عن العنوان الوطني",
                "ربط مع مركز المعلومات الوطني",
                "بوابة الجهة الخارجية",
                "البريد السعودي - سبل",
                "خارجي",
                "XML {ID}",
                "XML {Name}",
                "منصة التكامل الموحدة",
                "GSB",
                "Web Services",
                "XML",
            ],
        ],
    )
    return wb


def test_applications_workbook(tmp_path):
    snapshot = parse_xlsx_path(_save(_applications_workbook(), tmp_path))
    ents = _by_id(snapshot)

    app = ents["nora:Application:بوابه الجهه الخارجيه"]
    attrs = app.custom_fields
    assert attrs["appLayer"] == "access"
    assert attrs["numberOfUsers"] == 12000
    assert attrs["developmentType"] == "cots"
    assert attrs["sourceType"] == "inHouse"
    assert attrs["contractor"] == "شركة التقنية"
    assert attrs["appUrl"] == "http://www.example.gov.sa"
    assert attrs["launchDate"] == "2021-12-12"
    assert attrs["businessCriticality"] == "businessCritical"
    assert attrs["architecturePattern"] == "nTier"
    assert attrs["costCapex"] == 1000000
    assert attrs["costTotalAnnual"] == 500000
    # Active status + launch date → lifecycle.
    assert app.lifecycle == {"active": "2021-12-12"}

    iface = ents["nora:Interface:استعلام عن العنوان الوطني"]
    assert iface.custom_fields["integrationScope"] == "external"
    assert iface.custom_fields["linkType"] == "gsb"
    assert iface.custom_fields["viaGsb"] is True
    assert iface.custom_fields["protocol"] == "Web Services"
    assert iface.custom_fields["dataFormat"] == "XML"
    assert iface.custom_fields["integrationPlatform"] == "منصة التكامل الموحدة"

    # Consumer resolves to the in-file application, producer to a stub.
    triples = {(r.type, r.from_entity_id, r.to_entity_id) for r in snapshot.relations}
    assert ("applicationToInterface", app.source_id, iface.source_id) in triples
    assert (
        "applicationToInterface",
        "nora:Application:البريد السعودي - سبل",
        iface.source_id,
    ) in triples
    assert ents["nora:Application:البريد السعودي - سبل"].raw.get("nora_stub") is True


# ---------------------------------------------------------------------------
# Technology + Security workbooks
# ---------------------------------------------------------------------------


def test_technology_and_security_sheets(tmp_path):
    wb = Workbook()
    _write_sheet(
        wb,
        "أجهزة المضيف المادي",
        [
            ["دليل أجهزة المضيف المادي"],
            [
                "اسم المضيف",
                "الشركة المصنعة",
                "الطراز",
                "معرف الكتلة",
                "قطاع الشبكة",
                "تاريخ نهاية الدعم حسب المزود",
                "حالة عقد الدعم مع المزود",
                "نوع التشغيل",
                "التكلفة الأولية",
                "التكلفة التشغيلية",
            ],
            ["شرح"] * 10,
            [
                "RYD-Host01",
                "DELL",
                "Tier 3",
                "Cluster01",
                "DMZ",
                "31-12-2028",
                "Active",
                "Internal Team",
                165000,
                50000,
            ],
        ],
    )
    _write_sheet(
        wb,
        "دليل الخوادم الافتراضية",
        [
            ["دليل الخوادم الافتراضية"],
            [
                "اسم المضيف",
                "معرّف الكتلة",
                "قطاع الشبكة",
                "البيئة",
                "النسخ الاحتياطي",
                "التعافي من الكوارث",
                "نوع التشغيل",
                "التكلفة الأولية",
                "التكلفة التشغيلية",
            ],
            ["شرح"] * 9,
            ["RYDVM0081", "RYDCluster01", "DMZ", "Production", "Yes", "No", "Internal Team", 1, 2],
        ],
    )
    _write_sheet(
        wb,
        "برمجيات الأمن",
        [
            ["دليل برمجيات الأمن"],
            [
                "الرمز",
                "الاسم",
                "الشركة المصنعة",
                "الإصدار",
                "تاريخ نهاية الدعم حسب المزود",
                "حالة عقد الدعم مع المزود",
                "نوع التشغيل",
                "التكلفة الأولية",
                "التكلفة التشغيلية",
            ],
            ["شرح"] * 9,
            [
                "SecuritySoftware007",
                "Data Leakage Prevention (DLP)",
                "Symantec",
                "DLP 15.5",
                "31-12-2028",
                "expire",
                "Internal Team",
                750000,
                1750000,
            ],
        ],
    )
    snapshot = parse_xlsx_path(_save(wb, tmp_path))
    ents = _by_id(snapshot)

    host = ents["nora:PhysicalHost:ryd-host01"]
    assert host.category == "physicalHost"
    assert host.custom_fields["manufacturer"] == "DELL"
    assert host.custom_fields["modelNumber"] == "Tier 3"
    assert host.custom_fields["clusterId"] == "Cluster01"
    assert host.custom_fields["securityZone"] == "DMZ"
    assert host.custom_fields["supportEndDate"] == "2028-12-31"
    assert host.custom_fields["supportContractStatus"] == "active"
    assert host.custom_fields["operationType"] == "internalTeam"
    assert host.custom_fields["initialCost"] == 165000
    assert host.custom_fields["costTotalAnnual"] == 50000

    vm = ents["nora:VirtualServer:rydvm0081"]
    assert vm.category == "virtualServer"
    assert vm.custom_fields["environment"] == "production"
    assert vm.custom_fields["inBackupPolicy"] is True
    assert vm.custom_fields["inDrPolicy"] is False

    dlp = ents["nora:SecuritySoftware:data leakage prevention (dlp)"]
    assert dlp.category == "securitySoftware"
    assert dlp.custom_fields["supportContractStatus"] == "expired"
    assert dlp.custom_fields["version"] == "DLP 15.5"


# ---------------------------------------------------------------------------
# Data + Experience workbooks
# ---------------------------------------------------------------------------


def test_data_dictionary_sheet(tmp_path):
    wb = Workbook()
    _write_sheet(
        wb,
        "دليل قاموس البيانات",
        [
            ["دليل قاموس البيانات"],
            [
                "معرف كيان البيانات",
                "اسم كيان البيانات (انجليزي)",
                "اسم كيان البيانات(عربي)",
                "وصف الكيان (عربي)",
                "وصف الكيان (انجليزي)",
                "نوع البيانات",
                "تصنيف البيانات",
            ],
            ["شرح"] * 7,
            [
                "ENT_1",
                "Employee",
                "موظف",
                "بيانات موظفي الجهة",
                "Employee data",
                "Alphabetical",
                "مقيد",
            ],
        ],
    )
    snapshot = parse_xlsx_path(_save(wb, tmp_path))
    ents = _by_id(snapshot)
    obj = ents["nora:DataObject:موظف"]
    assert obj.type == "DataObject"
    assert obj.name == "موظف"
    assert obj.description == "بيانات موظفي الجهة"
    assert obj.custom_fields["dataType"] == "Alphabetical"
    assert obj.custom_fields["dataClassification"] == "restricted"
    assert obj.custom_fields["__name_en"] == "Employee"


def test_experience_workbook_is_descoped(tmp_path):
    wb = Workbook()
    _write_sheet(
        wb,
        " التحسينات على رحلات المس",
        [
            ["قائمة التحسينات على رحلات المستفيدين"],
            ["رحلة المستفيد", "المرحلة", "الفجوة"],
            ["شرح"] * 3,
            ["رحلة التقديم على شكوى", "المتابعة", "لا يمكن الاطلاع على الحالة"],
        ],
    )
    snapshot = parse_xlsx_path(_save(wb, tmp_path))
    assert snapshot.entities == []
    assert any("التحسينات" in e for e in snapshot.parse_errors)


# ---------------------------------------------------------------------------
# Adapter contract
# ---------------------------------------------------------------------------


def test_nora_source_is_registered():
    assert "nora_xlsx" in SOURCES
    src = SOURCES["nora_xlsx"]
    assert src.label == "NORA data-collection templates (DGA)"
    assert src.accepted_extensions == (".xlsx",)
    assert src.relation_mapping["serviceToApplication"] == "relGovServiceToApp"
    assert src.relation_mapping["applicationToInterface"] == "relAppToInterface"


def test_is_xlsx_payload():
    assert is_xlsx_payload(b"PK\x03\x04rest")
    assert not is_xlsx_payload(b"<?xml")


def test_stub_payload_never_blanks_description():
    """The post-build hook drops the empty description on stub entities so an
    update against an existing card can't wipe it."""
    from app.services.migration.snapshot import SourceEntity
    from app.services.migration.staging import build_card_payload

    src = SOURCES["nora_xlsx"]
    stub = SourceEntity(
        source_id="nora:Application:x",
        type="Application",
        name="X",
        raw={"nora_stub": True},
    )
    payload = build_card_payload(src, stub, "Application")
    assert "description" not in payload
    assert payload["attributes"]["source_origin"] == "nora_xlsx:referenced"

    real = SourceEntity(
        source_id="nora:Application:y", type="Application", name="Y", description="d"
    )
    payload = build_card_payload(src, real, "Application")
    assert payload["description"] == "d"
