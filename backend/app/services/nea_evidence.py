"""NEA evidence-pack aggregation + workbook builder ([FORK] — noraPlan.md WP5.3).

Reads the live landscape and NORA overlays and renders a multi-sheet ``.xlsx``:
Overview, EA Maturity, Program Status, BRM Coverage, Shared Services,
Standards Compliance, and Approval History. Each sheet builder degrades
gracefully if its source table is empty or a field is missing, so a pack can
be produced at any maturity of the deployment.
"""

from __future__ import annotations

import logging
from datetime import date, datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.card import Card
from app.models.event import Event
from app.models.maturity import MaturityAssessment, MaturityDimensionScore
from app.models.nora_program import NORA_STAGE_NUMBERS, EaProgramDeliverable
from app.models.tech_standard import StandardException, TechStandard
from app.models.user import User
from app.services.maturity import compute_overall_score

log = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14)

# Approval-related event types surfaced in the audit sheet.
_APPROVAL_EVENT_TYPES = (
    "card.approval_status_changed",
    "card.state_promoted",
    "nora_program.deliverable_updated",
    "maturity.assessment_updated",
)


def _truthy(v: object) -> bool:
    return v is True or (isinstance(v, str) and v.lower() in ("true", "yes", "1"))


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = _HEADER_FONT


def _autosize(ws: Worksheet) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(60, max(12, width + 2))


async def _gather(db: AsyncSession) -> dict:
    """Collect every dataset once. Sub-queries are individually guarded so a
    single failure never aborts the whole pack."""

    async def _safe(name: str, coro):
        try:
            return await coro
        except Exception as e:  # noqa: BLE001
            log.warning("nea evidence subsection %s failed: %s", name, e)
            try:
                await db.rollback()
            except Exception:
                pass
            return None

    data: dict = {}

    # --- User display names (for audit + generated-by) --------------------
    async def _users():
        rows = await db.execute(select(User.id, User.display_name))
        return {uid: name for uid, name in rows.all()}

    data["users"] = await _safe("users", _users()) or {}

    # --- Cards (non-archived) for BRM + shared services + landscape -------
    async def _cards():
        rows = await db.execute(
            select(
                Card.id,
                Card.type,
                Card.name,
                Card.attributes,
                Card.architecture_state,
                Card.approval_status,
            ).where(Card.status != "ARCHIVED")
        )
        return list(rows.all())

    data["cards"] = await _safe("cards", _cards()) or []

    # --- Program deliverables --------------------------------------------
    async def _deliverables():
        rows = await db.execute(
            select(EaProgramDeliverable).order_by(
                EaProgramDeliverable.stage_no, EaProgramDeliverable.sort_order
            )
        )
        return list(rows.scalars().all())

    data["deliverables"] = await _safe("deliverables", _deliverables()) or []

    # --- Latest maturity assessment + scores -----------------------------
    async def _maturity():
        latest = (
            await db.execute(
                select(MaturityAssessment)
                .order_by(MaturityAssessment.assessment_date.desc())
                .limit(1)
            )
        ).scalar_one_or_none()
        if latest is None:
            return None
        scores = list(
            (
                await db.execute(
                    select(MaturityDimensionScore).where(
                        MaturityDimensionScore.assessment_id == latest.id
                    )
                )
            )
            .scalars()
            .all()
        )
        return {"assessment": latest, "scores": scores}

    data["maturity"] = await _safe("maturity", _maturity())

    # --- Standards + exceptions ------------------------------------------
    async def _standards():
        standards = list((await db.execute(select(TechStandard))).scalars().all())
        exc_rows = await db.execute(
            select(
                StandardException.standard_id,
                StandardException.status,
                StandardException.expiry_date,
            )
        )
        return {"standards": standards, "exceptions": list(exc_rows.all())}

    data["standards"] = await _safe("standards", _standards()) or {
        "standards": [],
        "exceptions": [],
    }

    # --- Approval / governance audit events ------------------------------
    async def _events():
        rows = await db.execute(
            select(Event.event_type, Event.data, Event.user_id, Event.created_at, Event.card_id)
            .where(Event.event_type.in_(_APPROVAL_EVENT_TYPES))
            .order_by(Event.created_at.desc())
            .limit(500)
        )
        return list(rows.all())

    data["events"] = await _safe("events", _events()) or []

    return data


def _sheet_overview(wb: Workbook, data: dict, summary: dict) -> None:
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "NEA Alignment Evidence Pack"
    ws["A1"].font = _TITLE_FONT
    ws.append([])
    ws.append(["Generated at (UTC)", datetime.now(timezone.utc).isoformat(timespec="seconds")])
    ws.append([])
    _write_header(ws, ["Metric", "Value"])
    ws.append(["EA maturity — overall score (%)", summary.get("maturity_score")])
    ws.append(["EA program progress (%)", summary.get("program_progress")])
    ws.append(["BRM coverage (%)", summary.get("brm_coverage")])
    ws.append(["Business capabilities", summary.get("capabilities")])
    ws.append(["Shared-service candidates", summary.get("shared_services")])
    ws.append(["Technology standards", summary.get("standards")])
    ws.append(["Open standard exceptions", summary.get("open_exceptions")])
    ws.append(["Cards — current state", summary.get("state_current")])
    ws.append(["Cards — transition state", summary.get("state_transition")])
    ws.append(["Cards — target state", summary.get("state_target")])
    ws.append(["Approved cards", summary.get("approved_cards")])
    _autosize(ws)


def _sheet_maturity(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("EA Maturity")
    m = data.get("maturity")
    if not m:
        ws["A1"] = "No maturity assessment recorded."
        return
    a = m["assessment"]
    ws.append(["Assessment", a.title])
    ws.append(["Date", a.assessment_date.isoformat() if a.assessment_date else ""])
    ws.append(["Status", a.status])
    ws.append(["Overall score (%)", a.overall_score if a.overall_score is not None else ""])
    ws.append([])
    _write_header(ws, ["Dimension", "Current level", "Target level", "Gap"])
    for s in sorted(m["scores"], key=lambda s: (s.sort_order, s.dimension_name)):
        gap = max(0, (s.target_level or 0) - (s.level or 0))
        ws.append([s.dimension_name, s.level, s.target_level, gap])
    _autosize(ws)


def _sheet_program(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Program Status")
    deliverables = data.get("deliverables", [])
    _write_header(ws, ["Stage", "Deliverable", "Status"])
    by_stage: dict[int, list] = {n: [] for n in NORA_STAGE_NUMBERS}
    for d in deliverables:
        by_stage.setdefault(d.stage_no, []).append(d)
    for stage_no in sorted(by_stage):
        for d in by_stage[stage_no]:
            ws.append([stage_no, d.title, d.status])
    _autosize(ws)


def _sheet_brm(wb: Workbook, data: dict) -> tuple[int, float]:
    ws = wb.create_sheet("BRM Coverage")
    _write_header(ws, ["Capability", "BRM code", "BRM level"])
    caps = [c for c in data.get("cards", []) if c.type == "BusinessCapability"]
    mapped = 0
    for c in caps:
        attrs = c.attributes or {}
        code = attrs.get("brmCode") or ""
        level = attrs.get("brmLevel") or ""
        if code:
            mapped += 1
        ws.append([c.name, code, level])
    coverage = round(100 * mapped / len(caps), 1) if caps else 0.0
    ws.append([])
    ws.append(["Coverage (%)", coverage])
    ws.append(["Mapped / total", f"{mapped} / {len(caps)}"])
    _autosize(ws)
    return len(caps), coverage


def _sheet_shared_services(wb: Workbook, data: dict) -> int:
    ws = wb.create_sheet("Shared Services")
    _write_header(ws, ["Card", "Type", "Shared flag"])
    count = 0
    for c in data.get("cards", []):
        if c.type not in ("GovService", "Application"):
            continue
        attrs = c.attributes or {}
        shared = attrs.get("sharedService") or attrs.get("sharedServiceConsumer")
        if _truthy(shared):
            count += 1
            ws.append([c.name, c.type, "Yes"])
    if count == 0:
        ws.append(["No shared-service candidates flagged.", "", ""])
    _autosize(ws)
    return count


def _sheet_standards(wb: Workbook, data: dict) -> tuple[int, int]:
    ws = wb.create_sheet("Standards Compliance")
    std = data.get("standards", {})
    standards = std.get("standards", [])
    exceptions = std.get("exceptions", [])
    today = date.today()
    open_exc_by_std: dict = {}
    open_total = 0
    for standard_id, status, expiry in exceptions:
        active = status == "approved" and (expiry is None or expiry >= today)
        if status == "requested" or active:
            open_exc_by_std[standard_id] = open_exc_by_std.get(standard_id, 0) + 1
            open_total += 1
    _write_header(ws, ["Standard", "Status", "Mandate", "Issuing body", "Open exceptions"])
    for s in standards:
        ws.append(
            [s.name, s.status, s.mandate, s.standard_body or "", open_exc_by_std.get(s.id, 0)]
        )
    if not standards:
        ws.append(["No technology standards recorded.", "", "", "", ""])
    _autosize(ws)
    return len(standards), open_total


def _sheet_audit(wb: Workbook, data: dict) -> None:
    ws = wb.create_sheet("Approval History")
    _write_header(ws, ["Timestamp (UTC)", "Event", "User", "Detail"])
    users = data.get("users", {})
    for event_type, payload, user_id, created_at, _card_id in data.get("events", []):
        detail = ""
        if isinstance(payload, dict):
            detail = payload.get("status") or payload.get("action") or ""
        ws.append(
            [
                created_at.isoformat(timespec="seconds") if created_at else "",
                event_type,
                users.get(user_id, ""),
                detail,
            ]
        )
    _autosize(ws)


async def build_evidence_pack(db: AsyncSession) -> tuple[bytes, dict]:
    """Build the workbook, returning ``(xlsx_bytes, summary)``."""
    data = await _gather(db)

    # Landscape/approval roll-ups for the Overview + summary.
    cards = data.get("cards", [])
    state_counts = {"current": 0, "transition": 0, "target": 0}
    approved = 0
    for c in cards:
        st = c.architecture_state or "current"
        state_counts[st] = state_counts.get(st, 0) + 1
        if c.approval_status == "APPROVED":
            approved += 1

    deliverables = data.get("deliverables", [])
    in_scope = [d for d in deliverables if d.status != "descoped"]
    program_progress = (
        round(100 * len([d for d in in_scope if d.status == "approved"]) / len(in_scope))
        if in_scope
        else 0
    )

    maturity_score = None
    if data.get("maturity"):
        a = data["maturity"]["assessment"]
        maturity_score = (
            a.overall_score
            if a.overall_score is not None
            else compute_overall_score(data["maturity"]["scores"])
        )

    wb = Workbook()
    _sheet_maturity(wb, data)  # created after active; overview fills active below
    caps, brm_coverage = _sheet_brm(wb, data)
    shared = _sheet_shared_services(wb, data)
    n_standards, open_exc = _sheet_standards(wb, data)
    _sheet_program(wb, data)
    _sheet_audit(wb, data)

    summary = {
        "maturity_score": maturity_score,
        "program_progress": program_progress,
        "brm_coverage": brm_coverage,
        "capabilities": caps,
        "shared_services": shared,
        "standards": n_standards,
        "open_exceptions": open_exc,
        "state_current": state_counts.get("current", 0),
        "state_transition": state_counts.get("transition", 0),
        "state_target": state_counts.get("target", 0),
        "approved_cards": approved,
    }
    # Overview reuses the default active sheet, so it is already first.
    _sheet_overview(wb, data, summary)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), summary
