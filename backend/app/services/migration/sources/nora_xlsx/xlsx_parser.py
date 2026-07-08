"""Parser for the NORA/DGA حصر البيانات data-collection workbooks.

[FORK FEATURE] — noraPlan.md WP6.6.

Each of the six official DGA templates is one ``.xlsx`` workbook per EA
domain. All share the same layout: row 1 = sheet title, row 2 = column
headers (Arabic), row 3 = per-column explanations, rows 4+ = data, plus a
``Lookup``/``Lookups`` sheet holding the sanctioned option values.

The parser recognises sheets by normalized Arabic-token containment (robust
against whitespace/hamza variance between template revisions), maps columns
to Turbo EA attribute keys via per-sheet specs, and emits a source-neutral
:class:`~app.services.migration.snapshot.MigrationSnapshot`.

Identity: the templates carry no machine ids, so ``source_id`` is
synthesized from the normalized entity name (``nora:<NativeType>:<name>``).
That makes name-based cross-references (services → applications, procedures
→ services, integration points → applications) resolve naturally — within
one workbook, across workbooks imported earlier (via the identity map), and
against pre-existing Turbo EA cards (via the staging pipeline's
name-and-type fallback). Referenced names with no entity row become minimal
**stub** entities so the relation always lands; a stub that matches an
existing card by name simply binds to it.
"""

from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.migration.snapshot import MigrationSnapshot, Relation, SourceEntity
from app.services.migration.sources.nora_xlsx import mappings

# ---------------------------------------------------------------------------
# Normalization
# ---------------------------------------------------------------------------

_ARABIC_DIACRITICS = re.compile(r"[ؐ-ًؚ-ٰٟۖ-ۭـ]")
_WS = re.compile(r"\s+")


def _fold(value: Any) -> str:
    """Normalize for matching: trim, collapse whitespace, lowercase latin,
    strip Arabic diacritics/tatweel, fold hamza/teh-marbuta variants."""
    s = str(value or "")
    s = _ARABIC_DIACRITICS.sub("", s)
    s = s.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    s = s.replace("ة", "ه").replace("ى", "ي")
    s = _WS.sub(" ", s).strip().lower()
    return s


def _norm_name(value: Any) -> str:
    """Display-name normalization for ids: trim + collapse whitespace only
    (the visible name keeps its exact spelling)."""
    return _WS.sub(" ", str(value or "")).strip()


def is_xlsx_payload(prefix: bytes) -> bool:
    return prefix[:4] == b"PK\x03\x04"


def _match_option(value: Any, mapping: dict[str, str]) -> str | None:
    """Longest-token containment match of a cell value against a lookup map."""
    folded = _fold(value)
    if not folded:
        return None
    for token, key in sorted(mapping.items(), key=lambda kv: -len(kv[0])):
        if _fold(token) in folded:
            return key
    return None


def _match_bool(value: Any) -> bool | None:
    folded = _fold(value)
    if not folded:
        return None
    for token, flag in sorted(mappings.YES_NO.items(), key=lambda kv: -len(kv[0])):
        if _fold(token) == folded or folded.startswith(_fold(token)):
            return flag
    return None


_NUM_RE = re.compile(r"[\d][\d,.٫٬]*")


def _num(value: Any) -> float | int | None:
    if isinstance(value, (int, float)):
        return value
    m = _NUM_RE.search(str(value or ""))
    if not m:
        return None
    raw = m.group(0).replace(",", "").replace("٬", "").replace("٫", ".")
    try:
        f = float(raw)
    except ValueError:
        return None
    return int(f) if f.is_integer() else f


_DATE_DMY = re.compile(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$")
_DATE_YMD = re.compile(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})")


def _date(value: Any) -> str | None:
    """Coerce a cell to ``YYYY-MM-DD``. Handles datetime objects (openpyxl),
    ``31-12-2028`` (template convention) and ISO-ish strings."""
    if value is None or value == "":
        return None
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    m = _DATE_YMD.match(s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    m = _DATE_DMY.match(s)
    if m:
        return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    return None


_SPLIT_RE = re.compile(r"[\n\r،;؛|]+")
_ITEM_PREFIX = re.compile(r"^[\s•·▪‣*\-–—]*(?:\d+\s*[-.)ـ]\s*)?")


def _split_multi(value: Any) -> list[str]:
    """Split a multi-value cell (newline / Arabic comma / semicolon lists,
    with optional bullet or ``1-`` enumeration prefixes) into clean items."""
    out: list[str] = []
    for part in _SPLIT_RE.split(str(value or "")):
        item = _ITEM_PREFIX.sub("", part).strip()
        if item:
            out.append(_norm_name(item))
    return out


# ---------------------------------------------------------------------------
# Sheet specifications
#
# Each spec: native type, whether the sheet feeds an ITComponent subtype,
# and column rules matched by folded-header containment (first hit wins;
# longer tokens listed first where prefixes collide).
#
# Column rule kinds:
#   name / description / attr / option / multi_option / number / cost /
#   date / bool / parent_name / rel_targets / rel_sources /
#   beneficiary_group / lifecycle_status / skip
# ---------------------------------------------------------------------------

_SHEET_SPECS: list[dict[str, Any]] = [
    {
        "sheet_token": "دليل الخدمات",
        "native_type": "GovService",
        "columns": [
            ("رمز الخدمه", ("attr", "serviceCode")),
            ("اسم الخدمه", ("name",)),
            ("وصف الخدمه", ("description",)),
            ("تصنيف الخدمه", ("option", "serviceClassification", mappings.SERVICE_CLASSIFICATION)),
            ("الخدمه الاساسيه", ("parent_name",)),
            ("نوع الخدمه", ("option", "serviceType", mappings.SERVICE_TYPE)),
            ("مستوي اتمته الخدمه", ("option", "automationLevel", mappings.AUTOMATION_LEVEL)),
            (
                "التطبيقات المستخدمه",
                ("rel_targets", "serviceToApplication", "Application"),
            ),
            ("خطوات تنفيذ الخدمه", ("attr", "executionSteps")),
            ("التغطيه الجغرافيه", ("attr", "geoCoverage")),
            ("متطلبات وشروط", ("attr", "serviceRequirements")),
            ("قنوات تقديم الخدمه", ("multi_option", "deliveryChannel", mappings.DELIVERY_CHANNEL)),
            (
                "نوع المستفيد",
                (
                    "beneficiary_group",
                    "beneficiaryType",
                    {"افراد": "citizen", "قطاع اعمال": "business", "جهات حكوميه": "government"},
                ),
            ),
            ("مستوي نضج الخدمه", ("option", "serviceMaturity", mappings.SERVICE_MATURITY)),
            ("مدخلات الخدمه", ("attr", "serviceInputs")),
            ("مخرجات الخدمه", ("attr", "serviceOutputs")),
            ("وجود رسوم", ("option", "feeModel", mappings.FEE_MODEL)),
            ("الجهات المشاركه", ("attr", "participatingEntities")),
            ("معدل طلبات الخدمه", ("number", "monthlyTransactions")),
            ("اتفاقيه مستوي الخدمه", ("number", "slaDays")),
        ],
    },
    {
        "sheet_token": "دليل الاجراءات",
        "native_type": "BusinessProcess",
        "columns": [
            ("اسم الاجراء", ("name",)),
            ("وصف الاجراء", ("description",)),
            (
                "تصنيف الاجراء",
                ("option", "processClassification", mappings.PROCESS_CLASSIFICATION),
            ),
            ("مستوي اتمته الاجراء", ("option", "automationLevel", mappings.AUTOMATION_LEVEL)),
            (
                "الانظمه المستخدمه",
                ("rel_targets", "processToApplication", "Application"),
            ),
            ("مدخلات الاجراء", ("attr", "processInputs")),
            ("مخرجات الاجراء", ("attr", "processOutputs")),
            ("خدمات ذات علاقه", ("rel_sources", "serviceToProcess", "GovService")),
            ("مده تنفيذ", ("number", "durationDays")),
            ("حدث الاطلاق", ("attr", "triggerEvent")),
            ("قواعد العمل", ("attr", "businessRules")),
        ],
    },
    {
        "sheet_token": "سجل التطبيقات",
        "native_type": "Application",
        "columns": [
            ("عدد المستخدمين", ("number", "numberOfUsers")),
            ("الاسم", ("name",)),
            ("الوصف", ("description",)),
            ("الطبقه", ("option", "appLayer", mappings.APP_LAYER)),
            ("نوع التطوير", ("option", "developmentType", mappings.DEVELOPMENT_TYPE)),
            ("نوع المصدر", ("option", "sourceType", mappings.SOURCE_TYPE)),
            ("المقاول", ("attr", "contractor")),
            ("رابط التطبيق", ("attr", "appUrl")),
            ("نوع المصادقه", ("attr", "authenticationMethod")),
            ("تاريخ الاطلاق", ("date", "launchDate")),
            ("الاهميه", ("option", "businessCriticality", mappings.CRITICALITY)),
            ("حاله التطبيق", ("lifecycle_status",)),
            ("نمط الهيكليه", ("option", "architecturePattern", mappings.ARCHITECTURE_PATTERN)),
            ("التكلفه الراسماليه", ("cost", "costCapex")),
            ("التكلفه التشغيليه", ("cost", "costTotalAnnual")),
        ],
    },
    {
        "sheet_token": "نقاط الربط",
        "native_type": "Interface",
        "columns": [
            ("الرقم", ("attr", "integrationPointCode")),
            ("الاسم", ("name",)),
            ("الوصف", ("description",)),
            ("المستهلك", ("rel_sources", "applicationToInterface", "Application")),
            ("المنتج", ("rel_sources", "applicationToInterface", "Application")),
            ("نطاق الربط", ("option", "integrationScope", mappings.INTEGRATION_SCOPE)),
            ("المدخلات", ("attr", "interfaceInputs")),
            ("المخرجات", ("attr", "interfaceOutputs")),
            ("منصه التكامل", ("attr", "integrationPlatform")),
            ("نوع الربط", ("option", "linkType", mappings.LINK_TYPE)),
            ("طريقه الربط", ("attr", "protocol")),
            ("صيغه البيانات", ("attr", "dataFormat")),
        ],
    },
    {
        "sheet_token": "قاموس البيانات",
        "native_type": "DataObject",
        "columns": [
            # Arabic name preferred as the card name; the English name is
            # matched first (longer token) and kept as a fallback.
            ("اسم كيان البيانات (انجليزي", ("attr", "__name_en")),
            ("اسم كيان البيانات", ("name",)),
            ("وصف الكيان (انجليزي", ("attr", "__desc_en")),
            ("وصف الكيان", ("description",)),
            ("نوع البيانات", ("attr", "dataType")),
            (
                "تصنيف البيانات",
                ("option", "dataClassification", mappings.DATA_CLASSIFICATION),
            ),
        ],
    },
    # ---- Technology architecture (بنية التقنية) ----
    {
        "sheet_token": "المضيف المادي",
        "native_type": "PhysicalHost",
        "columns": [
            ("اسم المضيف", ("name",)),
            ("الشركه المصنعه", ("attr", "manufacturer")),
            ("الطراز", ("attr", "modelNumber")),
            ("معرف الكتله", ("attr", "clusterId")),
            ("قطاع الشبكه", ("attr", "securityZone")),
            ("تاريخ نهايه الدعم", ("date", "supportEndDate")),
            (
                "حاله عقد الدعم",
                ("option", "supportContractStatus", mappings.SUPPORT_CONTRACT_STATUS),
            ),
            ("نوع التشغيل", ("option", "operationType", mappings.OPERATION_TYPE)),
            ("التكلفه الاوليه", ("cost", "initialCost")),
            ("التكلفه التشغيليه", ("cost", "costTotalAnnual")),
        ],
    },
    {
        "sheet_token": "الخوادم الافتراضيه",
        "native_type": "VirtualServer",
        "columns": [
            ("اسم المضيف", ("name",)),
            ("معرف الكتله", ("attr", "clusterId")),
            ("قطاع الشبكه", ("attr", "securityZone")),
            ("البيئه", ("option", "environment", mappings.ENVIRONMENT)),
            ("النسخ الاحتياطي", ("bool", "inBackupPolicy")),
            ("التعافي من الكوارث", ("bool", "inDrPolicy")),
            ("نوع التشغيل", ("option", "operationType", mappings.OPERATION_TYPE)),
            ("التكلفه الاوليه", ("cost", "initialCost")),
            ("التكلفه التشغيليه", ("cost", "costTotalAnnual")),
        ],
    },
    {
        "sheet_token": "الاجهزه الشبكيه",
        "native_type": "NetworkDevice",
        "columns": [
            ("اسم المضيف", ("name",)),
            ("الشركه المصنعه", ("attr", "manufacturer")),
            ("الطراز", ("attr", "modelNumber")),
            ("جزء الشبكه", ("attr", "securityZone")),
            ("الوظيفه", ("attr", "deviceFunction")),
            ("اصدار البرنامج الثابت", ("attr", "firmwareVersion")),
            ("تاريخ انتهاء الدعم", ("date", "supportEndDate")),
            (
                "حاله عقد الدعم",
                ("option", "supportContractStatus", mappings.SUPPORT_CONTRACT_STATUS),
            ),
            ("نوع التشغيل", ("option", "operationType", mappings.OPERATION_TYPE)),
            ("التكلفه الاوليه", ("cost", "initialCost")),
            ("التكلفه التشغيليه", ("cost", "costTotalAnnual")),
        ],
    },
    {
        "sheet_token": "دليل التخزين",
        "native_type": "Storage",
        "columns": [
            ("اسم المضيف", ("name",)),
            ("الشركه المصنعه", ("attr", "manufacturer")),
            ("الطراز", ("attr", "modelNumber")),
            ("اصدار البرنامج الثابت", ("attr", "firmwareVersion")),
            ("جزء الشبكه", ("attr", "securityZone")),
            ("تاريخ نهايه الدعم", ("date", "supportEndDate")),
            (
                "حاله عقد الدعم",
                ("option", "supportContractStatus", mappings.SUPPORT_CONTRACT_STATUS),
            ),
            ("نوع التشغيل", ("option", "operationType", mappings.OPERATION_TYPE)),
            ("التكلفه الاوليه", ("cost", "initialCost")),
            ("التكلفه التشغيليه", ("cost", "costTotalAnnual")),
        ],
    },
    {
        "sheet_token": "ادوات التقنيه التحتيه",
        "native_type": "InfraTool",
        "columns": [
            ("اسم المضيف", ("name",)),
            ("الشركه المصنعه", ("attr", "manufacturer")),
            ("الاصدار", ("attr", "version")),
            ("الوظيفه", ("attr", "deviceFunction")),
            ("تاريخ نهايه الدعم", ("date", "supportEndDate")),
            (
                "حاله عقد الدعم",
                ("option", "supportContractStatus", mappings.SUPPORT_CONTRACT_STATUS),
            ),
            ("نوع التشغيل", ("option", "operationType", mappings.OPERATION_TYPE)),
            ("التكلفه الاوليه", ("cost", "initialCost")),
            ("التكلفه التشغيليه", ("cost", "costTotalAnnual")),
        ],
    },
    {
        "sheet_token": "خدمات التقنيه التحتيه",
        "native_type": "InfraService",
        "columns": [
            ("رمز الخدمه", ("attr", "serviceCode")),
            ("اسم المضيف", ("name",)),
            ("الشركه المصنعه", ("attr", "manufacturer")),
            ("الاصدار", ("attr", "version")),
            ("الوظيفه", ("attr", "deviceFunction")),
            ("تاريخ نهايه الدعم", ("date", "supportEndDate")),
            (
                "حاله عقد الدعم",
                ("option", "supportContractStatus", mappings.SUPPORT_CONTRACT_STATUS),
            ),
            ("نوع التشغيل", ("option", "operationType", mappings.OPERATION_TYPE)),
            ("التكلفه الاوليه", ("cost", "initialCost")),
            ("التكلفه التشغيليه", ("cost", "costTotalAnnual")),
        ],
    },
    # ---- Security architecture (بنية الأمن) ----
    {
        "sheet_token": "اجهزه الامن",
        "native_type": "SecurityHardware",
        "columns": [
            ("اسم المضيف", ("name",)),
            ("الشركه المصنعه", ("attr", "manufacturer")),
            ("الطراز", ("attr", "modelNumber")),
            ("جزء الشبكه", ("attr", "securityZone")),
            ("الوظيفه", ("attr", "deviceFunction")),
            ("اصدار البرنامج الثابت", ("attr", "firmwareVersion")),
            ("تاريخ انتهاء الدعم", ("date", "supportEndDate")),
            (
                "حاله عقد الدعم",
                ("option", "supportContractStatus", mappings.SUPPORT_CONTRACT_STATUS),
            ),
            ("نوع التشغيل", ("option", "operationType", mappings.OPERATION_TYPE)),
            ("التكلفه الاوليه", ("cost", "initialCost")),
            ("التكلفه التشغيليه", ("cost", "costTotalAnnual")),
        ],
    },
    {
        "sheet_token": "برمجيات الامن",
        "native_type": "SecuritySoftware",
        "columns": [
            ("الاسم", ("name",)),
            ("الشركه المصنعه", ("attr", "manufacturer")),
            ("الاصدار", ("attr", "version")),
            ("تاريخ نهايه الدعم", ("date", "supportEndDate")),
            (
                "حاله عقد الدعم",
                ("option", "supportContractStatus", mappings.SUPPORT_CONTRACT_STATUS),
            ),
            ("نوع التشغيل", ("option", "operationType", mappings.OPERATION_TYPE)),
            ("التكلفه الاوليه", ("cost", "initialCost")),
            ("التكلفه التشغيليه", ("cost", "costTotalAnnual")),
        ],
    },
    {
        "sheet_token": "خدمات الامن",
        "native_type": "SecurityService",
        "columns": [
            ("اسم الخدمه", ("name",)),
            ("الوظيفه", ("attr", "deviceFunction")),
            ("الوصف", ("description",)),
            (
                "حاله عقد الدعم",
                ("option", "supportContractStatus", mappings.SUPPORT_CONTRACT_STATUS),
            ),
            ("نوع التشغيل", ("option", "operationType", mappings.OPERATION_TYPE)),
            ("التكلفه الاوليه", ("cost", "initialCost")),
            ("التكلفه التشغيليه", ("cost", "costTotalAnnual")),
        ],
    },
]

# Sheets we recognise but deliberately do not import in v1 (see the
# documented descopes in noraPlan.md WP6.6).
_SKIPPED_SHEET_TOKENS: list[str] = [
    "المقدمه",
    "lookup",
    "دليل النماذج",
    "دليل السياسات",
    "اصحاب المصلحه",
    "فهرس الاعمال",
    "سمات البيانات",
    "التحسينات علي رحلات",
]


def _entity_id(native_type: str, name: str) -> str:
    return f"nora:{native_type}:{_fold(name)}"


def _resolve_headers(header_row: tuple, explanation_row: tuple) -> list[str]:
    """Effective header per column: row-2 header, falling back to the row-3
    explanation when the header cell is empty (the templates use merged
    group headers whose sub-columns are only labelled in the explanation
    row — e.g. the beneficiary-type yes/no triple)."""
    n = max(len(header_row), len(explanation_row))
    out: list[str] = []
    for i in range(n):
        h = header_row[i] if i < len(header_row) else None
        e = explanation_row[i] if i < len(explanation_row) else None
        out.append(_norm_name(h) or _norm_name(e))
    return out


def parse_xlsx_path(path: str | Path) -> MigrationSnapshot:
    data = Path(path).read_bytes()
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)

    entities: dict[str, SourceEntity] = {}
    relations: list[Relation] = []
    stub_names: dict[str, tuple[str, str]] = {}  # id -> (native_type, display name)
    parse_errors: list[str] = []

    for ws in wb.worksheets:
        title_folded = _fold(ws.title)
        spec = next((s for s in _SHEET_SPECS if _fold(s["sheet_token"]) in title_folded), None)
        if spec is None:
            if any(_fold(tok) in title_folded for tok in _SKIPPED_SHEET_TOKENS):
                parse_errors.append(f"Sheet '{ws.title.strip()}' is not imported (v1 descope)")
            else:
                parse_errors.append(f"Unrecognised sheet '{ws.title.strip()}' skipped")
            continue
        _parse_sheet(ws, spec, entities, relations, stub_names, parse_errors)

    wb.close()

    # Materialise stubs for referenced names that have no entity row —
    # relations must always have both endpoints in the snapshot.
    for stub_id, (native_type, display) in stub_names.items():
        if stub_id in entities:
            continue
        entities[stub_id] = SourceEntity(
            source_id=stub_id,
            type=native_type,
            name=display,
            raw={"nora_stub": True},
        )

    return MigrationSnapshot(
        version="nora-templates-1",
        entities=list(entities.values()),
        relations=relations,
        subscriptions=[],
        tags=[],
        documents=[],
        comments=[],
        users=[],
        metamodel_types=[],
        metamodel_relation_types=[],
        parse_errors=parse_errors,
    )


def _parse_sheet(
    ws,
    spec: dict[str, Any],
    entities: dict[str, SourceEntity],
    relations: list[Relation],
    stub_names: dict[str, tuple[str, str]],
    parse_errors: list[str],
) -> None:
    native_type = spec["native_type"]
    rows = ws.iter_rows(values_only=True)
    try:
        next(rows)  # row 1 — sheet title
        header_row = next(rows)
        explanation_row = next(rows)
    except StopIteration:
        parse_errors.append(f"Sheet '{ws.title.strip()}' has no data rows")
        return

    headers = _resolve_headers(header_row, explanation_row)

    # Resolve each column to a rule. Group columns (beneficiary triple)
    # match by the row-3 explanation once the row-2 header ran out.
    col_rules: list[tuple[int, tuple]] = []
    group_rule: tuple | None = next(
        (r for _, r in spec["columns"] if r[0] == "beneficiary_group"), None
    )
    for idx, header in enumerate(headers):
        folded = _fold(header)
        if not folded:
            continue
        matched = False
        for token, rule in spec["columns"]:
            if rule[0] == "beneficiary_group":
                continue
            if _fold(token) in folded:
                col_rules.append((idx, rule))
                matched = True
                break
        if matched:
            continue
        if group_rule is not None:
            # The group's own column carries the group header; its sibling
            # columns carry only the row-3 explanation label (أفراد / قطاع
            # أعمال / جهات حكومية). Either way the explanation (already the
            # fallback header for empty cells) decides the flag.
            _, attr_key, flag_map = group_rule
            explanation = _fold(explanation_row[idx] if idx < len(explanation_row) else "")
            probe = explanation or folded
            for flag_token, option_key in flag_map.items():
                if _fold(flag_token) in probe:
                    col_rules.append((idx, ("beneficiary_flag", attr_key, option_key)))
                    break

    name_col = next((i for i, r in col_rules if r[0] == "name"), None)
    if name_col is None:
        parse_errors.append(f"Sheet '{ws.title.strip()}': no name column recognised")
        return

    for row in rows:
        if row is None:
            continue
        name = _norm_name(row[name_col] if name_col < len(row) else None)
        if not name:
            continue

        attrs: dict[str, Any] = {}
        description: str | None = None
        parent_name: str | None = None
        beneficiary: list[str] = []
        lifecycle_phase: str | None = None
        row_relations: list[tuple[str, str, str, str]] = []  # (rel, other_type, other_name, dir)

        for idx, rule in col_rules:
            value = row[idx] if idx < len(row) else None
            if value is None or str(value).strip() == "":
                continue
            kind = rule[0]
            if kind == "name":
                continue
            elif kind == "description":
                description = _norm_name(value)
            elif kind == "attr":
                attrs[rule[1]] = _norm_name(value)
            elif kind == "option":
                hit = _match_option(value, rule[2])
                if hit is not None:
                    attrs[rule[1]] = hit
            elif kind == "multi_option":
                hits: list[str] = []
                for item in _split_multi(value):
                    hit = _match_option(item, rule[2])
                    if hit is not None and hit not in hits:
                        hits.append(hit)
                if hits:
                    attrs[rule[1]] = hits
            elif kind == "number":
                n = _num(value)
                if n is not None:
                    attrs[rule[1]] = n
            elif kind == "cost":
                n = _num(value)
                if n is not None:
                    attrs[rule[1]] = n
            elif kind == "date":
                d = _date(value)
                if d is not None:
                    attrs[rule[1]] = d
            elif kind == "bool":
                b = _match_bool(value)
                if b is not None:
                    attrs[rule[1]] = b
            elif kind == "parent_name":
                parent_name = _norm_name(value)
            elif kind == "beneficiary_flag":
                if _match_bool(value) is True:
                    beneficiary.append(rule[2])
            elif kind == "lifecycle_status":
                lifecycle_phase = _match_option(value, mappings.APP_STATUS_LIFECYCLE)
            elif kind == "rel_targets":
                for other in _split_multi(value):
                    row_relations.append((rule[1], rule[2], other, "out"))
            elif kind == "rel_sources":
                for other in _split_multi(value):
                    row_relations.append((rule[1], rule[2], other, "in"))

        if beneficiary:
            attrs["beneficiaryType"] = beneficiary

        # GSB-routed integration points also flip the WP1.1 viaGsb flag.
        if attrs.get("linkType") == "gsb":
            attrs["viaGsb"] = True

        # Application status → lifecycle (the WP6.2 mapping decision). Only
        # `active` carries a usable date (the launch date); other phases have
        # no date column in the template so lifecycle stays empty for them.
        lifecycle: dict[str, str] = {}
        if lifecycle_phase == "active" and attrs.get("launchDate"):
            lifecycle["active"] = attrs["launchDate"]

        entity_id = _entity_id(native_type, name)
        parent_id: str | None = None
        if parent_name:
            parent_id = _entity_id(native_type, parent_name)
            stub_names.setdefault(parent_id, (native_type, parent_name))

        entity = SourceEntity(
            source_id=entity_id,
            type=native_type,
            name=name,
            category=mappings.ITC_SUBTYPE_BY_NATIVE_TYPE.get(native_type),
            description=description,
            lifecycle=lifecycle,
            parent_id=parent_id,
            custom_fields=attrs,
            raw={"sheet": _norm_name(ws.title)},
        )
        if entity_id in entities and not entities[entity_id].raw.get("nora_stub"):
            parse_errors.append(
                f"Sheet '{ws.title.strip()}': duplicate name '{name}' — later row wins"
            )
        entities[entity_id] = entity

        seen_rel_ids = {r.source_id for r in relations}
        for rel_name, other_type, other_name, direction in row_relations:
            other_id = _entity_id(other_type, other_name)
            stub_names.setdefault(other_id, (other_type, other_name))
            from_id, to_id = (entity_id, other_id) if direction == "out" else (other_id, entity_id)
            rel_id = f"{from_id}::{rel_name}::{to_id}"
            if rel_id in seen_rel_ids:
                continue
            seen_rel_ids.add(rel_id)
            relations.append(
                Relation(
                    source_id=rel_id,
                    type=rel_name,
                    from_entity_id=from_id,
                    to_entity_id=to_id,
                )
            )
