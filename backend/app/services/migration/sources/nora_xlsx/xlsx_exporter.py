"""Exporter for the NORA/DGA حصر البيانات data-collection workbooks.

[FORK FEATURE] — noraPlan.md WP6.6 (deferred half — the submission roundtrip).

The reverse of :mod:`xlsx_parser`: takes a populated Turbo EA landscape and
writes it back into the official DGA template layout so an agency can submit
its filled files or round-trip-validate an import.

**Single source of truth**: the exporter reuses the parser's ``_SHEET_SPECS``
column definitions. Import and export therefore never drift — adding a column
rule to the parser makes it appear in the export automatically. For each
column rule the exporter runs the *inverse* of the parser's coercion
(option-key → Arabic label, phase → status label, related-card ids → names,
etc.).

Layout mirrors the importer: row 1 = sheet title, row 2 = Arabic headers,
row 3 = (left blank on export — the parser tolerates a missing explanation
row), rows 4+ = data.
"""

from __future__ import annotations

import uuid
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.card import Card
from app.models.relation import Relation
from app.services.migration.sources.nora_xlsx import mappings
from app.services.migration.sources.nora_xlsx.xlsx_parser import _SHEET_SPECS

# Domain key → the native sheet types it contains (mirrors the six DGA
# workbooks). One workbook per domain.
DOMAIN_NATIVE_TYPES: dict[str, list[str]] = {
    "business": ["GovService", "BusinessProcess"],
    "applications": ["Application", "Interface"],
    "data": ["DataObject"],
    "technology": [
        "PhysicalHost",
        "VirtualServer",
        "NetworkDevice",
        "Storage",
        "InfraTool",
        "InfraService",
    ],
    "security": ["SecurityHardware", "SecuritySoftware", "SecurityService"],
}

_HEADER_FILL = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
_TITLE_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")


def _reverse_option(mapping: dict[str, str]) -> dict[str, str]:
    """Build {option_key: first-seen Arabic label} from an import lookup map.

    The parser's maps are {arabic_or_bilingual_label: option_key}; several
    labels can point at one key. We pick the first label per key (dicts keep
    insertion order) as the canonical export label.
    """
    out: dict[str, str] = {}
    for label, key in mapping.items():
        out.setdefault(key, label)
    return out


# Precompute reverse maps once.
_REV_APP_STATUS = _reverse_option(mappings.APP_STATUS_LIFECYCLE)


def _stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "نعم" if value else "لا"
    if isinstance(value, list):
        return "، ".join(str(v) for v in value)
    return str(value)


class NoraTemplateExporter:
    """Build a NORA DGA template workbook for one domain from the landscape."""

    def __init__(self, db: AsyncSession):
        self.db = db
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # drop the default empty sheet

    async def build(self, domain: str) -> bytes:
        native_types = DOMAIN_NATIVE_TYPES.get(domain)
        if native_types is None:
            raise ValueError(f"Unknown domain: {domain!r}")

        for spec in _SHEET_SPECS:
            if spec["native_type"] not in native_types:
                continue
            await self._add_spec_sheet(spec)

        # Guarantee at least one sheet so openpyxl can save.
        if not self.wb.sheetnames:
            self.wb.create_sheet("empty")

        buf = BytesIO()
        self.wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    async def _add_spec_sheet(self, spec: dict) -> None:
        native_type = spec["native_type"]
        card_type = mappings.TYPE_MAPPING.get(native_type, native_type)
        subtype = mappings.ITC_SUBTYPE_BY_NATIVE_TYPE.get(native_type)

        stmt = select(Card).where(Card.type == card_type, Card.status != "ARCHIVED")
        if subtype is not None:
            stmt = stmt.where(Card.subtype == subtype)
        cards = (await self.db.execute(stmt)).scalars().all()

        # Resolve parent + related card names for the rows we're about to emit.
        card_by_id = {c.id: c for c in cards}
        rel_names = await self._related_names(spec, [c.id for c in cards])

        ws = self.wb.create_sheet(_sheet_title(spec["sheet_token"]))
        headers = [rule_header for rule_header, _ in spec["columns"]]

        # Row 1: title (merged look, just the token in A1).
        ws.append([spec["sheet_token"]])
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].fill = _TITLE_FILL

        # Row 2: Arabic headers.
        ws.append(headers)
        for cell in ws[2]:
            if cell.value:
                cell.font = Font(bold=True)
                cell.fill = _HEADER_FILL
                cell.alignment = Alignment(wrap_text=True, horizontal="right")

        # Row 3: explanation row. Must be a physically-present row (openpyxl
        # drops a truly empty ``append([])``), because the parser consumes
        # exactly three header rows before the data — an absent row 3 would
        # make it swallow the first data row as the explanation.
        ws.append([""] * len(headers))

        # Rows 4+: data.
        for card in cards:
            ws.append(self._row_for_card(card, spec, card_by_id, rel_names))

        for idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 22
        ws.sheet_view.rightToLeft = True

    async def _related_names(
        self, spec: dict, card_ids: list[uuid.UUID]
    ) -> dict[uuid.UUID, dict[str, list[str]]]:
        """For each card, map ``rel_type`` → list of related card names, for the
        ``rel_targets`` / ``rel_sources`` column rules on this sheet."""
        rel_rules = [r for _, r in spec["columns"] if r[0] in ("rel_targets", "rel_sources")]
        if not rel_rules or not card_ids:
            return {}
        rel_types = {r[1] for r in rel_rules}
        # Map the adapter's native relation name → TEA relation type key.
        tea_types = {mappings.RELATION_MAPPING.get(rt, rt) for rt in rel_types}

        rows = (
            (
                await self.db.execute(
                    select(Relation).where(
                        Relation.type.in_(tea_types),
                        Relation.source_id.in_(card_ids) | Relation.target_id.in_(card_ids),
                    )
                )
            )
            .scalars()
            .all()
        )

        needed_ids: set[uuid.UUID] = set()
        for rel in rows:
            needed_ids.add(rel.source_id)
            needed_ids.add(rel.target_id)
        names = dict(
            (await self.db.execute(select(Card.id, Card.name).where(Card.id.in_(needed_ids)))).all()
        )

        # native rel name → TEA type, for grouping back per rule.
        result: dict[uuid.UUID, dict[str, list[str]]] = {}
        for rel in rows:
            for native_rt in rel_types:
                if mappings.RELATION_MAPPING.get(native_rt, native_rt) != rel.type:
                    continue
                for cid, other in (
                    (rel.source_id, rel.target_id),
                    (rel.target_id, rel.source_id),
                ):
                    if cid in card_ids and other in names:
                        result.setdefault(cid, {}).setdefault(native_rt, []).append(names[other])
        return result

    def _row_for_card(
        self,
        card: Card,
        spec: dict,
        card_by_id: dict[uuid.UUID, Card],
        rel_names: dict[uuid.UUID, dict[str, list[str]]],
    ) -> list[str]:
        attrs = card.attributes or {}
        row: list[str] = []
        for _header, rule in spec["columns"]:
            row.append(self._cell(card, rule, attrs, card_by_id, rel_names))
        return row

    def _cell(
        self,
        card: Card,
        rule: tuple,
        attrs: dict,
        card_by_id: dict[uuid.UUID, Card],
        rel_names: dict[uuid.UUID, dict[str, list[str]]],
    ) -> str:
        kind = rule[0]
        if kind == "name":
            return card.name or ""
        if kind == "description":
            return card.description or ""
        if kind in ("attr", "number", "cost", "date"):
            return _stringify(attrs.get(rule[1]))
        if kind == "bool":
            v = attrs.get(rule[1])
            return "نعم" if v is True else ("لا" if v is False else "")
        if kind == "option":
            rev = _reverse_option(rule[2])
            return rev.get(attrs.get(rule[1]), "")
        if kind == "multi_option":
            rev = _reverse_option(rule[2])
            vals = attrs.get(rule[1]) or []
            if isinstance(vals, str):
                vals = [vals]
            return "، ".join(rev.get(v, "") for v in vals if rev.get(v))
        if kind == "parent_name":
            parent = card_by_id.get(card.parent_id) if card.parent_id else None
            return parent.name if parent else ""
        if kind == "beneficiary_group":
            # Emit the beneficiary keys as a joined cell (the parser's yes/no
            # triple collapses to a single list on the card).
            rev = {v: k for k, v in rule[2].items()}
            vals = attrs.get(rule[1]) or []
            if isinstance(vals, str):
                vals = [vals]
            return "، ".join(rev.get(v, v) for v in vals)
        if kind == "lifecycle_status":
            phases = (card.lifecycle or {}).keys()
            for phase in phases:
                if phase in _REV_APP_STATUS:
                    return _REV_APP_STATUS[phase]
            return ""
        if kind in ("rel_targets", "rel_sources"):
            names = rel_names.get(card.id, {}).get(rule[1], [])
            return "، ".join(sorted(set(names)))
        return ""


def _sheet_title(token: str) -> str:
    """openpyxl caps sheet names at 31 chars and forbids some characters."""
    safe = token.replace("/", "-").replace("\\", "-").replace("*", "").strip()
    return safe[:31] or "sheet"


async def build_template_workbook(db: AsyncSession, domain: str) -> bytes:
    """Public entry point — build the DGA template workbook for one domain."""
    return await NoraTemplateExporter(db).build(domain)
