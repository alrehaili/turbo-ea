"""Reference Models helpers ([FORK] — noraPlan.md WP100.3).

Serialisers, the single-published-per-domain resolver, the xlsx workbook
builder/parser (Code | Parent Code | Name | Name (Arabic) | Description |
Sort Order), the code-upsert engine shared by both import endpoints, and the
kit-preview starter seed. Kept out of the route module so the pure parts are
unit-testable without a request.
"""

from __future__ import annotations

import uuid
from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.card import Card
from app.models.reference_model import (
    ReferenceModel,
    ReferenceModelItem,
    ReferenceModelMapping,
    ReferenceModelVersion,
)

_HEADER_FONT = Font(bold=True)

# Domain → (card type key, card-attribute code field) that carries the RM
# classification of inventory cards. The same wiring backs the card-detail
# code pickers and the Reference Models report; browse pages count mapped
# inventory through it.
RM_DOMAIN_CARD_TYPES: dict[str, tuple[str, str]] = {
    "business": ("BusinessCapability", "brmCode"),
    "beneficiaryExperience": ("GovService", "bxrmCode"),
    "applications": ("Application", "armCode"),
    "data": ("DataObject", "drmCode"),
    "technology": ("ITComponent", "trmCode"),
    "security": ("SecurityControl", "srmCode"),
}

RM_EXPORT_HEADERS = ["Code", "Parent Code", "Name", "Name (Arabic)", "Description", "Sort Order"]

# Kit-preview starters ("عينة توضيحية" — illustrative, from the DGA awareness
# kit's National-RM preview slides). Seeded as draft/national/built_in so an
# agency reviews and publishes deliberately; content is data, admin-editable.
# (key, domain, name, name_ar, description, items[(code, name, name_ar)])
STARTER_REFERENCE_MODELS: tuple[dict, ...] = (
    {
        "key": "nea_business_preview",
        "domain": "business",
        "name": "National Business RM (kit preview)",
        "name_ar": "النموذج المرجعي الوطني للأعمال (معاينة)",
        "description": (
            "Capability categories from the DGA awareness-kit preview of the National "
            "Business Reference Model. Illustrative — review against the published RM "
            "document before publishing."
        ),
        "items": (
            ("BRM-1", "Administrative & Regulatory Capabilities", "القدرات الإدارية والتنظيمية"),
            ("BRM-2", "Core Capabilities", "القدرات الأساسية"),
            ("BRM-3", "Supporting Capabilities", "القدرات المساندة"),
            ("BRM-4", "Operational Capabilities", "القدرات التشغيلية"),
            ("BRM-5", "Enabling Capabilities", "القدرات التمكينية"),
        ),
    },
    {
        "key": "nea_applications_preview",
        "domain": "applications",
        "name": "National Applications RM (kit preview)",
        "name_ar": "النموذج المرجعي الوطني للتطبيقات (معاينة)",
        "description": (
            "Application layers from the DGA awareness-kit preview of the National "
            "Applications Reference Model (same scheme as the appLayer profile field). "
            "Illustrative — review against the published RM document before publishing."
        ),
        "items": (
            ("ARM-1", "Access Layer", "طبقة الوصول"),
            ("ARM-2", "Core Layer", "الطبقة الأساسية"),
            ("ARM-3", "Support Layer", "طبقة المساندة"),
            ("ARM-4", "Data Layer", "طبقة البيانات"),
            ("ARM-5", "Infrastructure Layer", "طبقة البنية التحتية"),
        ),
    },
)


# --------------------------------------------------------------------------- #
# Serialisers
# --------------------------------------------------------------------------- #
def model_dict(m: ReferenceModel, item_count: int | None = None, names: dict | None = None) -> dict:
    names = names or {}
    out = {
        "id": str(m.id),
        "domain": m.domain,
        "key": m.key,
        "name": m.name,
        "name_ar": m.name_ar,
        "description": m.description,
        "version": m.version,
        "source": m.source,
        "status": m.status,
        "built_in": m.built_in,
        "created_by": str(m.created_by) if m.created_by else None,
        "published_by": str(m.published_by) if m.published_by else None,
        "published_by_display_name": names.get(m.published_by),
        "published_at": m.published_at.isoformat() if m.published_at else None,
        "created_at": m.created_at.isoformat() if m.created_at else None,
        "narrative": m.narrative or {"panels": []},
    }
    if item_count is not None:
        out["item_count"] = item_count
    return out


def item_dict(i: ReferenceModelItem) -> dict:
    return {
        "id": str(i.id),
        "model_id": str(i.model_id),
        "parent_id": str(i.parent_id) if i.parent_id else None,
        "code": i.code,
        "name": i.name,
        "name_ar": i.name_ar,
        "description": i.description,
        "sort_order": i.sort_order,
    }


def sort_items(items: list[ReferenceModelItem]) -> list[ReferenceModelItem]:
    return sorted(items, key=lambda i: (i.sort_order, i.code))


# --------------------------------------------------------------------------- #
# Queries
# --------------------------------------------------------------------------- #
async def resolve_active_model(db: AsyncSession, domain: str) -> ReferenceModel | None:
    """The single published RM for a domain (publish supersedes, so ≤1)."""
    return (
        (
            await db.execute(
                select(ReferenceModel)
                .where(ReferenceModel.domain == domain, ReferenceModel.status == "published")
                .order_by(ReferenceModel.published_at.desc().nulls_last())
                .limit(1)
            )
        )
        .scalars()
        .first()
    )


async def items_for(db: AsyncSession, model_id: uuid.UUID) -> list[ReferenceModelItem]:
    return list(
        (
            await db.execute(
                select(ReferenceModelItem).where(ReferenceModelItem.model_id == model_id)
            )
        )
        .scalars()
        .all()
    )


async def item_counts(db: AsyncSession, model_ids: list[uuid.UUID]) -> dict[uuid.UUID, int]:
    if not model_ids:
        return {}
    res = await db.execute(
        select(ReferenceModelItem.model_id, func.count())
        .where(ReferenceModelItem.model_id.in_(model_ids))
        .group_by(ReferenceModelItem.model_id)
    )
    return dict(res.all())


async def scan_domain_card_codes(
    db: AsyncSession, domain: str
) -> list[tuple[uuid.UUID, str | None]]:
    """``(card_id, code_or_None)`` for every non-archived card of the domain type."""
    wiring = RM_DOMAIN_CARD_TYPES.get(domain)
    if wiring is None:
        return []
    card_type, code_field = wiring
    rows = await db.execute(
        select(Card.id, Card.attributes).where(Card.type == card_type, Card.status != "ARCHIVED")
    )
    out: list[tuple[uuid.UUID, str | None]] = []
    for cid, attributes in rows.all():
        code = (attributes or {}).get(code_field)
        code = code.strip() if isinstance(code, str) and code.strip() else None
        out.append((cid, code))
    return out


async def explicit_mappings_for_model(
    db: AsyncSession, model_id: uuid.UUID
) -> list[ReferenceModelMapping]:
    return list(
        (
            await db.execute(
                select(ReferenceModelMapping).where(ReferenceModelMapping.model_id == model_id)
            )
        )
        .scalars()
        .all()
    )


def subtree_id_sets(
    items: list[ReferenceModelItem], direct: dict[uuid.UUID, set[uuid.UUID]]
) -> dict[uuid.UUID, tuple[set[uuid.UUID], set[uuid.UUID]]]:
    """Per-item ``(own_card_ids, descendant-inclusive_card_ids)``.

    ``direct`` maps *item id* → set of directly-mapped card ids. Cycle-safe.
    """
    children: dict[uuid.UUID | None, list[ReferenceModelItem]] = {}
    known = {i.id for i in items}
    for item in items:
        parent = item.parent_id if item.parent_id in known else None
        children.setdefault(parent, []).append(item)

    out: dict[uuid.UUID, tuple[set[uuid.UUID], set[uuid.UUID]]] = {}

    def walk(item: ReferenceModelItem, seen: set[uuid.UUID]) -> set[uuid.UUID]:
        if item.id in seen:
            return set()
        seen.add(item.id)
        own = set(direct.get(item.id, set()))
        total = set(own)
        for child in children.get(item.id, []):
            total |= walk(child, seen)
        out[item.id] = (own, total)
        return total

    for root in children.get(None, []):
        walk(root, set())
    for item in items:  # defensive: any node a cycle excluded still gets a row
        d = set(direct.get(item.id, set()))
        out.setdefault(item.id, (d, set(d)))
    return out


def compute_model_coverage(
    items: list[ReferenceModelItem],
    card_codes: list[tuple[uuid.UUID, str | None]],
    mappings: list[ReferenceModelMapping],
) -> dict:
    """Merge code-attribute classification with explicit mappings (dedup by card).

    A card maps to an item if its code equals ``item.code`` OR a non-rejected
    explicit mapping row links the two. Model-level KPIs partition the domain's
    cards into mapped / unmatched (a code that resolves to no item) / uncoded.
    Returns ``{per_item, total_cards, mapped_cards, unmatched_cards,
    uncoded_cards, covered_items}``.
    """
    item_codes = {i.code for i in items}
    code_ids: dict[str, set[uuid.UUID]] = {}
    for cid, code in card_codes:
        if code:
            code_ids.setdefault(code, set()).add(cid)

    direct: dict[uuid.UUID, set[uuid.UUID]] = {
        i.id: set(code_ids.get(i.code, set())) for i in items
    }
    explicit_card_ids: set[uuid.UUID] = set()
    for m in mappings:
        if m.mapping_status != "rejected":
            direct.setdefault(m.item_id, set()).add(m.card_id)
            explicit_card_ids.add(m.card_id)

    per_item = subtree_id_sets(items, direct)
    covered_items = sum(1 for _own, total in per_item.values() if total)

    mapped = {
        cid for cid, code in card_codes if (code and code in item_codes) or cid in explicit_card_ids
    }
    unmatched = {cid for cid, code in card_codes if cid not in mapped and code}
    uncoded = {cid for cid, code in card_codes if cid not in mapped and not code}

    return {
        "per_item": per_item,
        "total_cards": len(card_codes),
        "mapped_cards": len(mapped),
        "unmatched_cards": len(unmatched),
        "uncoded_cards": len(uncoded),
        "covered_items": covered_items,
    }


_RETIRING_PHASES = {"phaseOut", "endOfLife"}


def current_lifecycle_phase(lifecycle: dict | None) -> str | None:
    """The card's current lifecycle phase from its dated phases (latest ≤ today)."""
    if not lifecycle:
        return None
    today = datetime.now(timezone.utc).date()
    for phase in ("endOfLife", "phaseOut", "active", "phaseIn", "plan"):
        raw = lifecycle.get(phase)
        if not raw:
            continue
        try:
            d = (
                datetime.fromisoformat(str(raw)).date()
                if "T" in str(raw)
                else datetime.strptime(str(raw), "%Y-%m-%d").date()
            )
        except (ValueError, TypeError):
            continue
        if d <= today:
            return phase
    for phase in ("plan", "phaseIn", "active", "phaseOut", "endOfLife"):
        if lifecycle.get(phase):
            return phase
    return None


def is_retiring(lifecycle: dict | None) -> bool:
    return current_lifecycle_phase(lifecycle) in _RETIRING_PHASES


def leaf_item_ids(items: list[ReferenceModelItem]) -> set[uuid.UUID]:
    """Item ids that have no children (the mappable capabilities)."""
    parents = {i.parent_id for i in items if i.parent_id is not None}
    return {i.id for i in items if i.id not in parents}


def snapshot_items(items: list[ReferenceModelItem]) -> list[dict]:
    """Freeze the item tree into a code-keyed, diff-able snapshot list."""
    by_id = {i.id: i for i in items}
    out = []
    for i in sort_items(items):
        parent = by_id.get(i.parent_id) if i.parent_id else None
        out.append(
            {
                "code": i.code,
                "parent_code": parent.code if parent else None,
                "name": i.name,
                "name_ar": i.name_ar,
                "description": i.description,
                "sort_order": i.sort_order,
            }
        )
    return out


_DIFF_FIELDS = ("name", "name_ar", "description", "parent_code", "sort_order")


def diff_snapshots(before: list[dict], after: list[dict]) -> dict:
    """Diff two snapshots by code → added / removed / changed / unchanged."""
    a = {r["code"]: r for r in (before or [])}
    b = {r["code"]: r for r in (after or [])}
    added = [b[c] for c in b if c not in a]
    removed = [a[c] for c in a if c not in b]
    changed = []
    unchanged = 0
    for code in a.keys() & b.keys():
        fields = [f for f in _DIFF_FIELDS if a[code].get(f) != b[code].get(f)]
        if fields:
            changed.append({"code": code, "before": a[code], "after": b[code], "fields": fields})
        else:
            unchanged += 1
    added.sort(key=lambda r: r["code"])
    removed.sort(key=lambda r: r["code"])
    changed.sort(key=lambda r: r["code"])
    return {
        "added": added,
        "removed": removed,
        "changed": changed,
        "unchanged": unchanged,
        "counts": {
            "added": len(added),
            "removed": len(removed),
            "changed": len(changed),
            "unchanged": unchanged,
        },
    }


def version_dict(
    v: ReferenceModelVersion, names: dict | None = None, include_snapshot: bool = False
) -> dict:
    names = names or {}
    out = {
        "id": str(v.id),
        "model_id": str(v.model_id),
        "version": v.version,
        "change_summary": v.change_summary,
        "item_count": v.item_count,
        "published_by": str(v.published_by) if v.published_by else None,
        "published_by_display_name": names.get(v.published_by),
        "published_at": v.published_at.isoformat() if v.published_at else None,
        "created_at": v.created_at.isoformat() if v.created_at else None,
    }
    if include_snapshot:
        out["snapshot"] = v.snapshot or []
    return out


async def versions_for(db: AsyncSession, model_id: uuid.UUID) -> list[ReferenceModelVersion]:
    return list(
        (
            await db.execute(
                select(ReferenceModelVersion)
                .where(ReferenceModelVersion.model_id == model_id)
                .order_by(ReferenceModelVersion.published_at.desc().nulls_last())
            )
        )
        .scalars()
        .all()
    )


def mapping_dict(m: ReferenceModelMapping, names: dict | None = None) -> dict:
    names = names or {}
    return {
        "id": str(m.id),
        "model_id": str(m.model_id),
        "item_id": str(m.item_id),
        "card_id": str(m.card_id),
        "mapping_type": m.mapping_type,
        "mapping_status": m.mapping_status,
        "rationale": m.rationale,
        "confidence": m.confidence,
        "reviewed_at": m.reviewed_at.isoformat() if m.reviewed_at else None,
        "reviewed_by": str(m.reviewed_by) if m.reviewed_by else None,
        "reviewed_by_display_name": names.get(m.reviewed_by),
        "created_by": str(m.created_by) if m.created_by else None,
        "created_at": m.created_at.isoformat() if m.created_at else None,
    }


# --------------------------------------------------------------------------- #
# xlsx build / parse
# --------------------------------------------------------------------------- #
def build_rm_workbook(model: ReferenceModel, items: list[ReferenceModelItem]) -> bytes:
    """One-sheet workbook in the canonical RM exchange layout."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reference Model"
    ws.append(RM_EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = _HEADER_FONT

    by_id = {i.id: i for i in items}
    for i in sort_items(items):
        parent = by_id.get(i.parent_id) if i.parent_id else None
        ws.append(
            [
                i.code,
                parent.code if parent else None,
                i.name,
                i.name_ar,
                i.description,
                i.sort_order,
            ]
        )

    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(60, max(12, width + 2))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _match_header(value: object) -> str | None:
    h = str(value or "").strip().lower()
    if not h:
        return None
    if "parent" in h or "أصل" in h or "الأب" in h:
        return "parent_code"
    if "arabic" in h or "عرب" in h:
        return "name_ar"
    if h == "code" or "code" in h or "رمز" in h:
        return "code"
    if "name" in h or "اسم" in h or "الاسم" in h:
        return "name"
    if "desc" in h or "وصف" in h:
        return "description"
    if "sort" in h or "order" in h or "ترتيب" in h:
        return "sort_order"
    return None


def parse_rm_workbook(data: bytes) -> tuple[list[dict], list[str]]:
    """Parse the RM exchange layout into row dicts + human-readable errors.

    Tolerant header matching (English/Arabic), first sheet only, row 1 headers.
    Duplicate codes and unknown parent codes are reported as errors; valid rows
    still parse so a partial file lands with warnings rather than failing whole.
    """
    errors: list[str] = []
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return [], ["Not a readable .xlsx workbook"]

    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return [], ["The first sheet is empty"]

    col_map: dict[int, str] = {}
    for idx, value in enumerate(header_row):
        field = _match_header(value)
        if field and field not in col_map.values():
            col_map[idx] = field
    if "code" not in col_map.values() or "name" not in col_map.values():
        return [], ["Header row must contain at least Code and Name columns"]

    rows: list[dict] = []
    seen_codes: set[str] = set()
    for row_no, row in enumerate(rows_iter, start=2):
        raw: dict = {}
        for idx, field in col_map.items():
            raw[field] = row[idx] if idx < len(row) else None
        code = str(raw.get("code") or "").strip()
        name = str(raw.get("name") or "").strip()
        if not code and not name:
            continue  # blank spacer row
        if not code:
            errors.append(f"Row {row_no}: missing code — skipped")
            continue
        if not name:
            errors.append(f"Row {row_no}: missing name — skipped")
            continue
        if code in seen_codes:
            errors.append(f"Row {row_no}: duplicate code '{code}' — skipped")
            continue
        seen_codes.add(code)
        sort_raw = raw.get("sort_order")
        try:
            sort_order = int(sort_raw) if sort_raw not in (None, "") else 0
        except (TypeError, ValueError):
            sort_order = 0
        rows.append(
            {
                "code": code[:64],
                "parent_code": str(raw.get("parent_code") or "").strip()[:64] or None,
                "name": name[:255],
                "name_ar": (str(raw.get("name_ar") or "").strip() or None),
                "description": (str(raw.get("description") or "").strip() or None),
                "sort_order": sort_order,
            }
        )

    codes = {r["code"] for r in rows}
    for r in rows:
        if r["parent_code"] and r["parent_code"] not in codes:
            errors.append(
                f"Item '{r['code']}': parent code '{r['parent_code']}' not found — "
                "imported as a root item"
            )
            r["parent_code"] = None
    return rows, errors


async def upsert_items(db: AsyncSession, model: ReferenceModel, rows: list[dict]) -> dict[str, int]:
    """Upsert parsed rows into a model by code (two-pass parent resolution)."""
    existing = {i.code: i for i in await items_for(db, model.id)}
    created = updated = unchanged = 0

    for r in rows:
        item = existing.get(r["code"])
        if item is None:
            item = ReferenceModelItem(
                model_id=model.id,
                code=r["code"],
                name=r["name"],
                name_ar=r["name_ar"],
                description=r["description"],
                sort_order=r["sort_order"],
            )
            db.add(item)
            existing[r["code"]] = item
            created += 1
        elif (
            item.name != r["name"]
            or item.name_ar != r["name_ar"]
            or item.description != r["description"]
            or item.sort_order != r["sort_order"]
        ):
            item.name = r["name"]
            item.name_ar = r["name_ar"]
            item.description = r["description"]
            item.sort_order = r["sort_order"]
            updated += 1
        else:
            unchanged += 1
    await db.flush()

    parent_changed = 0
    for r in rows:
        item = existing[r["code"]]
        new_parent = existing[r["parent_code"]].id if r["parent_code"] else None
        if item.parent_id != new_parent:
            item.parent_id = new_parent
            parent_changed += 1
    await db.flush()

    return {
        "created": created,
        "updated": updated,
        "unchanged": unchanged,
        "parents_changed": parent_changed,
    }


# --------------------------------------------------------------------------- #
# Starter seed (called from the NORA profile apply)
# --------------------------------------------------------------------------- #
async def seed_reference_model_starters(db: AsyncSession) -> int:
    """Idempotently insert the kit-preview starter RMs (by key)."""
    existing = {
        k
        for (k,) in (
            await db.execute(select(ReferenceModel.key).where(ReferenceModel.key.is_not(None)))
        ).all()
    }
    created = 0
    for starter in STARTER_REFERENCE_MODELS:
        if starter["key"] in existing:
            continue
        model = ReferenceModel(
            key=starter["key"],
            domain=starter["domain"],
            name=starter["name"],
            name_ar=starter["name_ar"],
            description=starter["description"],
            version="preview",
            source="national",
            status="draft",
            built_in=True,
        )
        db.add(model)
        await db.flush()
        for sort_order, (code, name, name_ar) in enumerate(starter["items"]):
            db.add(
                ReferenceModelItem(
                    model_id=model.id,
                    code=code,
                    name=name,
                    name_ar=name_ar,
                    sort_order=sort_order,
                )
            )
        created += 1
    await db.flush()
    return created
